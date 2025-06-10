import numpy as np
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill
from models import BaseDataMapping, PfltSecurityMapping

import os
import traceback
import math

# Get the root directory (where the script is located)
root_path = os.path.dirname(os.path.abspath(__file__))

print("Root path:", root_path)


def extract_continuous_data(df, issuer_col, fmv_col):
    clean_data = []
    for i, val in enumerate(df[issuer_col]):
        if pd.isna(val) or str(val).strip() == "":
            break
        clean_data.append((val, df[fmv_col].iloc[i]))
    return pd.DataFrame(clean_data, columns=[issuer_col, fmv_col])

def extract_and_add_similar(file_name, sheet_name, mapping_columns, compare_columns):
    try:
        # Read raw Excel to find where header row starts
        df_raw = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
        issuer_position = df_raw.isin(mapping_columns)

        if issuer_position.any().any():
            row_index, col_index = issuer_position.stack()[lambda x: x].index[0]
            print(f"Issuer found at row {row_index}, column {col_index}")
        else:
            raise ValueError("Issuer column not found")

        # Read again with header row
        df = pd.read_excel(file_name, sheet_name=sheet_name, header=row_index)

        # specific to PCOF
        if "Investment" in df.columns:
            stop_index = df[df["Investment"] == "Total"].index.min()
            if pd.notna(stop_index):
                df = df.loc[:stop_index - 1]  # keep all rows before that row
        # if "Investment Name" in df.columns:
        #     security_mapping = PfltSecurityMapping.query.filter_by(pcof_report_inv_name is not None).all()


        for col_name in mapping_columns:
            if col_name not in df.columns:
                raise ValueError(col_name + " not present in the file")
        # Validate presence of columns
        for col_name in compare_columns:
            if col_name not in df.columns:
                # raise ValueError(col_name + " not present in the file")
                print(col_name + " not present in the file")

        # Step 2: Create final DataFrame preserving the order
        final_df = pd.DataFrame(df, columns = mapping_columns + compare_columns)
        return final_df
    except Exception as e:
        # traceback.print_exc()
        print(e)

def is_number(value):
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def compare_data(source_df, target_df, source_mapping_cols, target_mapping_cols, source_compare_cols, target_compare_cols):
    """
    Compare source and target dataframes based on mapping columns and compare columns.
    
    Returns a dataframe with mismatches.
    """
    
    diff_counts = []
    # Rename columns to avoid confusion during merge
    source_df = source_df.copy()
    target_df = target_df.copy()

    source_renames = {s: f"src_{s}" for s in source_compare_cols}
    target_renames = {t: f"tgt_{t}" for t in target_compare_cols}

    source_df.rename(columns=source_renames, inplace=True)
    target_df.rename(columns=target_renames, inplace=True)

    source_df = source_df.drop_duplicates(source_mapping_cols)
    target_df = target_df.drop_duplicates(target_mapping_cols)

    # Make lowercase copies of the mapping columns
    for col in source_mapping_cols:
        source_df[col] = source_df[col].str.lower() if source_df[col].dtype == 'object' else source_df[col]

    for col in target_mapping_cols:
        target_df[col] = target_df[col].str.lower() if target_df[col].dtype == 'object' else target_df[col]


    # Merge on mapping columns
    merged_df = source_df.merge(
        target_df,
        left_on=source_mapping_cols,
        right_on=target_mapping_cols,
        how='outer',  # capture all differences
        suffixes=('_src', '_tgt'),
        indicator=True
    )
    merged_df = merged_df.drop_duplicates()
    differences = pd.DataFrame()

    for src_col, tgt_col in zip(source_compare_cols, target_compare_cols):
        src_col_renamed = f"src_{src_col}"
        tgt_col_renamed = f"tgt_{tgt_col}"
        mismatch_mask = merged_df[src_col_renamed] != merged_df[tgt_col_renamed]
        # mismatch_rows = merged_df[mismatch_mask]

        if not merged_df.empty:
            mismatch_rows = merged_df[source_mapping_cols + [src_col_renamed, tgt_col_renamed]]

            diff_col = []
            diff_col_percent = []

            for idx, (src_val, tgt_val) in enumerate(zip(mismatch_rows[src_col_renamed], mismatch_rows[tgt_col_renamed])):
                try:
                    # Convert to float
                    if is_number(src_val) and is_number(tgt_val):
                        src_num = float(src_val)
                        tgt_num = float(tgt_val)
                        if (pd.isna(src_val) and not pd.isna(tgt_val) and tgt_num != 0) or (not pd.isna(src_val) and pd.isna(tgt_val) and src_num != 0):
                            diff_col.append("Yes")
                            diff_col_percent.append("")
                        # Round for comparison
                        elif round(src_num, 2) != round(tgt_num, 2):
                            diff = round(src_num - tgt_num, 3)
                            diff_col.append(diff)
                            if not pd.isna(tgt_val) and tgt_num != 0:
                                diff_col_percent.append(diff/tgt_num)
                            else:
                                diff_col_percent.append("")
                        else:
                            diff_col.append("")
                            diff_col_percent.append("")
                    else:
                        if (src_val is None and tgt_val is not None) or (src_val is not None and tgt_val is None)or (pd.isna(src_val) and not pd.isna(tgt_val)) or (not pd.isna(src_val) and pd.isna(tgt_val)) or src_val != tgt_val:
                            diff_col.append("Yes")
                            diff_col_percent.append("")
                        else:
                            diff_col.append("")
                            diff_col_percent.append("")
                except (ValueError, TypeError):
                    # Handle conversion errors or missing values
                    diff_col.append("")
                    diff_col_percent.append("")
            non_empty_count = sum(
                1 for x in diff_col if x is not None and str(x).strip() != "" and not (isinstance(x, float) and math.isnan(x))
            )

            zero_value_count = sum(
                1 for x in diff_col if x is not None and str(x).strip() != "" and is_number(x) and float(x) == 0.0
            )

            # Clean and convert to float where valid
            cleaned = []
            for val in diff_col:
                try:
                    if val is not None and str(val).strip() != "" and not (isinstance(val, float) and math.isnan(val)):
                        cleaned.append(float(val))
                except ValueError:
                    continue  # skip non-numeric values
            # Calculate mean
            mean_val = sum(cleaned) / len(cleaned) if cleaned else None
            median_val = np.median(cleaned) if cleaned else None
            std_dev = np.std(cleaned, ddof=0) if cleaned else None

            diff_counts.append([f"diff_{src_col}", non_empty_count, zero_value_count, mean_val, median_val, std_dev])
            mismatch_rows[f"diff_{src_col}"] = diff_col
            mismatch_rows[f"diff_{src_col}_percent"] = diff_col_percent
            if differences.empty:
                differences = mismatch_rows
            else:
                differences = differences.merge(
                    mismatch_rows,
                    left_on=source_mapping_cols,
                    right_on=source_mapping_cols,
                    how='outer',  # capture all differences
                    suffixes=('_src', '_tgt'),
                    indicator=True
                )
                if '_merge' in differences.columns:
                    differences = differences.drop(columns=['_merge'])
            # differences.append(merged_df)

    if not differences.empty:
        # result_df = pd.concat(differences, ignore_index=True)
        result_df = differences
        result_df = result_df.sort_values(by=source_mapping_cols[0])
    else:
        result_df = pd.DataFrame(columns=source_mapping_cols + ['column', 'source_value', 'target_value'])

    return result_df, diff_counts

sheets_to_compare =[
    ['PL BB Build', 'PL BB Build'],
    # ['InputPLBB', 'InputPLBB']
]
mapping_columns = {
    'PL BB Build': ['Issuer', 'Investment Name'],
    'InputPLBB': ['Investment Type']
}

# the mapping columns should not be in the below list
PCOF_columns = {
    'PL BB Build': [
    ["Leverage LTV Thru PCOF IV", "LTV Thru PCOF IV"],
    ["Financials LTM EBITDA ($MMs)", "LTM EBITDA ($MMs)"],
    ["Financials LTM Revenue ($MMs)", "LTM Revenue ($MMs)"],
    # ["Investment Name", "Investment Name"],  # commented in first list
    # ["Rates Fixed / Floating", "Fixed / Floating"],
    ["Rates Floating Cash Spread", "Floating Cash Spread"],
    ["Investment Industry", "Industry"],
    ["Rates Current LIBOR/Floor", "Current LIBOR/Floor"],
    # ["Investment Cost", "Cost"],  # commented
    # ["Investment External Valuation", "External Valuation"],  # commented
    # ["Leverage Attachment Point", "Attachment Point"],
    # ["Borrowing Base Industry Concentration", "Industry Concentration"],  # commented
    # ["Borrowing Base Comment", "Comment"],  # commented
    # ["Is Eligible Issuer", "Eligible"],  # commented
    # ["Classifications Defaulted / Restructured", "Defaulted / Restructured"],  # commented
    ["Leverage Total Capitalization", "Total Capitalization"],
    # ["Investment Closing Date", "Closing Date"],
    # ["Classifications Noteless Assigned Loan", "Noteless Assigned Loan"],
    ["Leverage PCOF IV Leverage", "PCOF IV Leverage"],
    # ["Investment Par", "Par"],  # commented
    # ["Classifications Quoted / Unquoted", "Quoted / Unquoted"],  # commented
    # ["Classifications Warehouse Asset", "Warehouse Asset"],  # commented
    # ["Classifications Warehouse Asset Expected Rating", "Warehouse Asset Expected Rating"],  # commented
    # ["Classifications Approved Foreign Jurisdiction", "Approved Foreign Jurisdiction"],  # commented
    # ["Classifications LTV Transaction", "LTV Transaction"],  # commented
    # ["Classifications Undelivered Note", "Undelivered Note"],  # commented
    # ["Classifications Structured Finance Obligation", "Structured Finance Obligation"],  # commented
    # ["Classifications Third Party Finance Company", "Third Party Finance Company"],  # commented
    # ["Classifications Affiliate Investment", "Affiliate Investment"],  # commented
    # ["Investment Internal Valuation", "Internal Valuation"],  # commented
    # ["Rates PIK", "PIK"],
    # ["Classifications Warehouse Asset Inclusion Date", "Warehouse Asset Inclusion Date"],  # commented
    # ["Leverage Revolver Commitment", "Revolver Commitment"],  # commented in first list, uncommented in second
    # ["Final Eligibility Override", "Eligibility Override"],  # commented
    # ["Final Comment", "Comment"],  # commented
    # ["Concentration Adjustment", "Adjustment"],  # commented
    # ["Concentration Comment", "Comment"],  # commented
    # ["Borrowing Base Other Adjustment", "Other Adjustment"],  # commented
    # ["Investment Maturity", "Maturity"],  # commented
    ["Leverage Total Enterprise Value", "Total Enterprise Value"],
    # ["Investment Investment Type", "Investment Type"],  # commented
    ["Leverage Total Leverage", "Total Leverage"],
    ["Rates Fixed Coupon", "Fixed Coupon"],
    # ["Adv. Adv. Rate", "Adv. Rate"],
    # ["Borrowing Base", "Borrowing Base"]
    ],
    'InputPLBB': [
        ['Unquoted', 'Unquoted'],
        ['Quoted', 'Quoted']
    ]
}

def compare_pcof_report():
    data = {}

    for sheet_name in sheets_to_compare:
        print(sheet_name)
        # pepper report
        source_data = {
            "file_path": "",
            "sheet_name": sheet_name[0],
            "mapping_columns": mapping_columns[sheet_name[0]],
            "compare_columns": [x[0] for x in PCOF_columns[sheet_name[0]]]
        }
        # client report
        target_data = {
            "file_path": "",
            "sheet_name": sheet_name[1],
            "mapping_columns": mapping_columns[sheet_name[1]],
            "compare_columns": [x[1] for x in PCOF_columns[sheet_name[0]]]
        }

        source_data_df = extract_and_add_similar(source_data['file_path'], source_data['sheet_name'], source_data['mapping_columns'], source_data['compare_columns'])
        target_data_df = extract_and_add_similar(target_data['file_path'], target_data['sheet_name'], target_data['mapping_columns'], target_data['compare_columns'])

        if source_data_df is None or target_data_df is None:
            print('No data found')
        else:
            # Merge on 'Issuer' to find common issuers
            comparison_df, diff_counts = compare_data(source_data_df, target_data_df, source_data['mapping_columns'], target_data['mapping_columns'], source_data['compare_columns'], target_data['compare_columns'])
            comparison_df = comparison_df.where(pd.notna(comparison_df), None)
            # Sort by column 'your_column' in ascending order
            comparison_df = comparison_df.sort_values(by=mapping_columns[sheet_name[0]], ascending=True)
            diff_df = pd.DataFrame(diff_counts, columns=["Columns", "Differences Count", "Zero Values", "Mean Difference", "Median Difference", "Standard Deviation of Diff"])
            columns_mapping_details = []
            mapping_details = BaseDataMapping.query.filter_by(fund_type='PCOF').all()
            for c in mapping_details:
                if c.bd_column_name in source_data['compare_columns']:
                    columns_mapping_details.append([c.bd_column_name, c.sf_sheet_name, c.sf_column_name])
            mapping_df = pd.DataFrame(columns_mapping_details, columns=["Column name", "Source sheet", "Source column"])
            data[sheet_name[0]] = [comparison_df, diff_df, mapping_df]

    file_path = "Comparison Report.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for key, value in data.items():
            value[0].to_excel(writer, sheet_name=key, index=False, header=True)
            value[1].to_excel(writer, sheet_name=key+"_diff_report", index=False, header=True)
            value[2].to_excel(writer, sheet_name="mapping details", index=False, header=True)
            # Load the workbook and sheet
            workbook = writer.book
            worksheet = writer.sheets[key]

            # Define a fill style (e.g., light red)
            highlight_fill = PatternFill(start_color="ffff86", end_color="ffff86", fill_type="solid")
            highlight_fill_red = PatternFill(start_color="fab6b6", end_color="fab6b6", fill_type="solid")

            # Get header
            header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            
            # Loop through rows and apply formatting
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for i, cell in enumerate(row):
                    if header[i].startswith('diff_') and not header[i].endswith('_percent'):
                        # if cell.column_letter == "A":  # Example: Highlight column A
                        if cell.value != "No" and cell.value is not None and cell.value != "":
                            cell.fill = highlight_fill
                    if header[i].startswith('diff_') and header[i].endswith('_percent'):
                        try:
                            # Convert value to float and format as percentage
                            val = float(cell.value)
                            cell.value = val
                            cell.number_format = "0.00%"
                            if val > 0.1 or val < -0.1:
                                cell.fill = highlight_fill_red
                        except (ValueError, TypeError):
                            pass  # Skip if not a number