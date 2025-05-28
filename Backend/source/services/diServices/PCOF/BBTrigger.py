import os
import pandas as pd
import numpy as np
import datetime
from sqlalchemy import text
import openpyxl
import pickle
import json

from models import db, BaseDataFile, ExtractedBaseDataInfo, BaseDataOtherInfo

from source.utility.ServiceResponse import ServiceResponse
from source.services.PCOF.PcofBBCalculator import PcofBBCalculator
from source.services.PCOF.PcofDashboardService import PcofDashboardService
from source.utility.Log import Log

pcofDashboardService = PcofDashboardService()
pcofBBCalculator = PcofBBCalculator()

def trigger_pcof_bb(bdi_id):
    try:
        engine = db.get_engine()
        extracted_base_data_info = ExtractedBaseDataInfo.query.filter_by(id=bdi_id).first()
        with engine.connect() as connection:
            base_data_df = pd.DataFrame(connection.execute(text(f'select * from pcof_base_data where base_data_info_id = :ebd_id'), {'ebd_id': bdi_id}).fetchall())
            base_data_mapping_df = pd.DataFrame(connection.execute(text("""select bd_sheet_name, bd_column_name, bd_column_lookup from base_data_mapping bdm where fund_type = 'PCOF'""")))
            pl_bb_result_security_df = pd.DataFrame(connection.execute(text('''select ps.security_name as "Security" from pcof_securities ps''')))
            input_industries_df = pd.DataFrame(connection.execute(text('''select distinct std_industry_name as "Industries" from pcof_industries pi''')))
            pl_bb_results = pd.DataFrame(connection.execute(text('''select ct.test_name as "Concentration Tests", fct.limit_percentage as "Concentration Limit" from concentration_test ct, fund_concentration_test fct where fct.fund_id = 1''')))


            base_data_df = base_data_df.replace({np.nan: None})
            report_date = base_data_df['report_date'][0].strftime("%Y-%m-%d")
            base_data_df = base_data_df.drop('report_date', axis=1)
            base_data_df = base_data_df.drop('created_at', axis=1)
            base_data_df = base_data_df.drop('base_data_info_id', axis=1)
            base_data_df = base_data_df.drop('id', axis=1)
            base_data_df = base_data_df.drop('company_id', axis=1)
            base_data_df = base_data_df.drop('created_by', axis=1)
            base_data_df = base_data_df.drop('modified_by', axis=1)
            base_data_df = base_data_df.drop('modified_at', axis=1)

        # base_data_df["Investment Internal Valuation"] = np.nan # complete column is empty
        # base_data_df["Rates PIK"] = np.nan # complete column is empty
        # base_data_df["Classifications Warehouse Asset Inclusion Date"] = np.nan # complete column is empty
        # base_data_df["Leverage Attachment Point"] = np.nan # complete column is empty
        # base_data_df["Final Eligibility Override"] = np.nan # complete column is empty
        # base_data_df["Final Comment"] = np.nan # complete column is empty
        # base_data_df["Concentration Comment"] = np.nan # complete column is empty
        # base_data_df["Borrowing Base Other Adjustment"] = np.nan # complete column is empty
        # base_data_df["Borrowing Base Industry Concentration"] = np.nan # complete column is empty
        # base_data_df["Borrowing Base Comment"] = np.nan # complete column is empty

        # base_data_df["Is Eligible Issuer"] = 'Yes'

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "PL BB Build"
        file_name = "PCOF Base data - " + report_date + ".xlsx"
        wb.save(file_name)

        path1 = 'PCOF Base data.xlsx'
        path2 = file_name
        wb1 = openpyxl.load_workbook(filename=path1)
        for index, sheet in enumerate(wb1.worksheets):
            ws1 = wb1.worksheets[index]

            wb2 = openpyxl.load_workbook(filename=path2)
            ws2 = wb2.create_sheet(ws1.title)
            for row in ws1:
                for cell in row:
                    ws2[cell.coordinate].value = cell.value
            wb2.save(path2)
        
        rename_df_col = {}
        for index, row in base_data_mapping_df.iterrows():
            rename_df_col[row['bd_column_lookup']] = row['bd_column_name']
        base_data_df.rename(columns=rename_df_col, inplace=True)

        xl_df_map = {}
        # xl_df_map['PL BB Build'] = base_data_df

        base_data_other_info = BaseDataOtherInfo.query.filter_by(extraction_info_id=bdi_id).first()

        # availability_borrower_data = [
        #     {"A": "Borrower:", "B": "PennantPark Credit Opportunities Fund IV"},
        #     {"A": "Date of determination:", "B": datetime.datetime(2024, 12, 31, 0, 0, 0)},
        #     {"A": "Revolving Closing Date", "B": datetime.datetime(2022, 12, 19, 0, 0, 0)},
        #     {"A": "Commitment Period (3 years from Final Closing Date, as defined in LPA):", "B": "Yes"},
        #     {"A": "(b) Facility Size:", "B":  240000000},
        #     {"A": "Loans (USD)", "B":  None},
        #     {"A": "Loans (CAD)", "B":  None}
        # ]
        availability_borrower_data = [
            {"A": "Borrower:", "B": base_data_other_info.other_info_list.get('availability_borrower').get('borrower')},
            {"A": "Date of determination:", "B": datetime.datetime.strptime(base_data_other_info.other_info_list.get('availability_borrower').get('determination_date')[:-5], "%Y-%m-%dT%H:%M:%S")},
            {"A": "Revolving Closing Date", "B": datetime.datetime.strptime(base_data_other_info.other_info_list.get('availability_borrower').get('revolving_closing_date')[:-5], "%Y-%m-%dT%H:%M:%S")},
            {"A": "Commitment Period (3 years from Final Closing Date, as defined in LPA)", "B": base_data_other_info.other_info_list.get('availability_borrower').get('commitment_period_(3_years_from_final_closing_date,_as_defined_in_lpa)')},
            {"A": "(b) Facility Size", "B":  base_data_other_info.other_info_list.get('availability_borrower').get('(b)_facility_size')},
            {"A": "Loans (USD)", "B":  base_data_other_info.other_info_list.get('availability_borrower').get('loans_(usd)')},
            {"A": "Loans (CAD)", "B":  base_data_other_info.other_info_list.get('availability_borrower').get('loans_(cad)')}
        ]
        df_Availability_Borrower = pd.DataFrame(availability_borrower_data)

        subscription_bb_data = [
            {
                "Investor": sub_bb_data.get('investor'),
                "Master/Feeder": sub_bb_data.get('master/feeder'),
                "Ultimate Investor Parent": sub_bb_data.get('ultimate_investor_parent'),
                "Designation": sub_bb_data.get('designation'),
                "Commitment": sub_bb_data.get('commitment'),
                "Capital Called": sub_bb_data.get('capital_called')
            } for sub_bb_data in base_data_other_info.other_info_list.get('subscription_bb')]

    
        df_subscription_bb = pd.DataFrame(subscription_bb_data)
        df_subscription_bb["Commitment"] = pd.to_numeric(df_subscription_bb["Commitment"], errors='coerce')
        df_subscription_bb["Capital Called"] = pd.to_numeric(df_subscription_bb["Capital Called"], errors='coerce')
        df_subscription_bb["Uncalled Capital"] = (df_subscription_bb["Commitment"] - df_subscription_bb["Capital Called"])
        total_capitalCalled = df_subscription_bb["Capital Called"].sum()
        total_uncalled_Capital = df_subscription_bb["Uncalled Capital"].sum()

        input_other_metrics_data = [{
                "Other Metrics": "First Lien Leverage Cut-Off Point",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('first_lien_leverage_cut-off_point')
            },
            {
                "Other Metrics": "Warehouse First Lien Leverage Cut-Off",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('warehouse_first_lien_leverage_cut-off')
            },
            {
                "Other Metrics": "Last Out Attachment Point",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('last_out_attachment_point')
            },
            {
                "Other Metrics": "1 out of 2 Test",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('1_out_of_2_test')
            },
            {
                "Other Metrics": "Trailing 12-Month EBITDA",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('trailing_12-month_ebitda')
            },
            {
                "Other Metrics": "Trailing 24-Month EBITDA",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('trailing_24-month_ebitda')
            },
            {
                "Other Metrics": "Total Leverage",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('total_leverage')
            },
            {
                "Other Metrics": "LTV",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('ltv')
            },
            {
                "Other Metrics": "Concentration Test Threshold 1",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('concentration_test_threshold_1')
            },
            {
                "Other Metrics": "Concentration Test Threshold 1",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('concentration_test_threshold_2')
            },
            {
                "Other Metrics": "Threshold 1 Advance Rate",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('threshold_1_advance_rate')
            },
            {
                "Other Metrics": "Threshold 2 Advance Rate",
                "values": base_data_other_info.other_info_list.get('other_metrics').get('threshold_2_advance_rate')
            }]
        df_Inputs_Other_Metrics = pd.DataFrame(input_other_metrics_data)

        
        inputs_portfolio_lev_bb_data = [{
                "Investment Type": plbb_data.get('investment_type'),
                "Unquoted": None if plbb_data.get('unquoted') == 'n/a' else plbb_data.get('unquoted'),
                "Quoted": plbb_data.get('quoted')
            } for plbb_data in base_data_other_info.other_info_list.get('portfolio_leverageborrowingbase')
        ]
        df_Inputs_Portfolio_LeverageBorrowingBase = pd.DataFrame(inputs_portfolio_lev_bb_data)

        concentration_limits_data = [{
                'Investors': cld.get('investors'),
                'Rates': cld.get('rates'),
                'Concentration Limit': cld.get('concentration_limit')
            } for cld in base_data_other_info.other_info_list.get('concentration_limits')
        ]
        concentration_limits_df = pd.DataFrame(concentration_limits_data)

        obligors_net_capital_data =[
            {
                "Obligors' Net Capital": onc.get("obligors_net_capital"),
                "Values": onc.get("values")
            }  for onc in base_data_other_info.other_info_list.get('obligors_net_capital')
        ]
        df_Obligors_Net_Capital = pd.DataFrame(obligors_net_capital_data)

        advance_rates_data = [{
                'Investor Type': ard.get('investor_type'),
                'Advance Rate': ard.get('advance_rate')
            } for ard in base_data_other_info.other_info_list.get('advance_rates')
        ]
        advance_rates_df = pd.DataFrame(advance_rates_data)

        principle_obligation_data = [{
            "Principal Obligations": pod.get('principal_obligations'),
            "Currency": pod.get('currency'),
            "Amount": pod.get('amount'),
            "Spot Rate": pod.get('spot_rate'),
            "Dollar Equivalent": pod.get('dollar_equivalent')
        } for pod in  base_data_other_info.other_info_list.get('principle_obligations')]
        principle_obligation_df = pd.DataFrame(principle_obligation_data)

        pricing_data = [{
            'Pricing': p_data.get('pricing'),
            'percent': p_data.get('percent')
        } for p_data in base_data_other_info.other_info_list.get('pricing')] 
        pricing_df = pd.DataFrame(pricing_data)     

        df_PL_BB_Build = base_data_df.copy()

        df_PL_BB_Build[["Classifications Structured Finance Obligation", 
                        "Classifications Third Party Finance Company", 
                        "Classifications Affiliate Investment", 
                        "Classifications Defaulted / Restructured"]].fillna('') # Note: This initialized to '', value must be asked.
        
        df_PL_BB_Build[["Final Eligibility Override"]].fillna('') # Note: This initialized to '', value must be asked.

        df_PL_BB_Build["Investment Cost"] = df_PL_BB_Build["Investment Cost"].astype(float)
        df_PL_BB_Build[["Investment Cost"]].fillna(0)
        df_PL_BB_Build["Investment Par"] = df_PL_BB_Build["Investment Cost"].astype(float)
        df_PL_BB_Build[["Investment Par"]].fillna(0)

        df_PL_BB_Build["Leverage PCOF IV Leverage"].fillna(0, inplace=True)

        df_PL_BB_Build["Rates Fixed Coupon"] = pd.to_numeric(df_PL_BB_Build["Rates Fixed Coupon"])
        df_PL_BB_Build["Rates Fixed Coupon"] = df_PL_BB_Build["Rates Fixed Coupon"].astype(float)
        df_PL_BB_Build["Rates Fixed Coupon"] = df_PL_BB_Build["Rates Fixed Coupon"].fillna(0).astype(float)

        df_PL_BB_Build["Rates Current LIBOR/Floor"] = pd.to_numeric(df_PL_BB_Build["Rates Current LIBOR/Floor"], errors='coerce')
        df_PL_BB_Build["Rates Current LIBOR/Floor"] = df_PL_BB_Build["Rates Current LIBOR/Floor"].fillna(0).astype(int)

        df_PL_BB_Build["Rates Floating Cash Spread"] = pd.to_numeric(df_PL_BB_Build["Rates Floating Cash Spread"], errors='coerce')
        df_PL_BB_Build["Rates Floating Cash Spread"] = df_PL_BB_Build["Rates Floating Cash Spread"].astype(float)
        df_PL_BB_Build["Rates Floating Cash Spread"] = df_PL_BB_Build["Rates Floating Cash Spread"].fillna(0)

        df_PL_BB_Build['Investment Maturity'] = pd.to_datetime(df_PL_BB_Build['Investment Maturity'], errors='coerce')
        
        df_PL_BB_Build['Investment Closing Date'] = pd.to_datetime(df_PL_BB_Build['Investment Closing Date'], errors='coerce')
        

        # df_PL_BB_Build = calculation_for_build(
        #     df_PL_BB_Build, 
        #     df_Inputs_Other_Metrics, 
        #     df_Availability_Borrower,
        #     total_capitalCalled,
        #     df_Inputs_Portfolio_LeverageBorrowingBase,
        #     total_uncalled_Capital,
        #     df_Obligors_Net_Capital
        # ) # For now, checking PL BB Build

        # base_data_df.to_excel(writer, sheet_name="PL BB Build", index=False, header=True)
        # writer.save()
        with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
            df_PL_BB_Build.to_excel(writer, sheet_name="PL BB Build", index=False, header=True)
            df_Inputs_Other_Metrics.to_excel(writer, sheet_name="Other Metrics", index=False, header=True)
            df_Availability_Borrower.to_excel(writer, sheet_name="Availability Borrower", index=False, header=True)
            df_Inputs_Portfolio_LeverageBorrowingBase.to_excel(writer, sheet_name="Portfolio LeverageBorrowingBase", index=False, header=True)
            df_Obligors_Net_Capital.to_excel(writer, sheet_name="Obligors' Net Capital", index=False, header=True)
            pl_bb_results.to_excel(writer, sheet_name="PL BB Results", index=False, header=True)
            df_subscription_bb.to_excel(writer, sheet_name="Subscription BB", index=False, header=True)
            pl_bb_result_security_df.to_excel(writer, sheet_name="PL_BB_Results_Security", index=False, header=True)
            input_industries_df.to_excel(writer, sheet_name="Inputs Industries", index=False, header=True)
            advance_rates_df.to_excel(writer, sheet_name="Advance Rates", index=False, header=True)
            concentration_limits_df.to_excel(writer, sheet_name="Concentration Limits", index=False, header=True)
            principle_obligation_df.to_excel(writer, sheet_name="Principle Obligations", index=False, header=True)
            pricing_df.to_excel(writer, sheet_name="Pricing", index=False, header=True)

        xl_df_map = pd.read_excel(file_name, sheet_name=["PL BB Build", "Other Metrics", "Availability Borrower", "Portfolio LeverageBorrowingBase", "Obligors' Net Capital", "PL BB Results", "Subscription BB", "PL_BB_Results_Security", "Inputs Industries", "Advance Rates", "Concentration Limits", "Principle Obligations", "Pricing"])
        # xl_df_map = {
        #     "PL BB Build": df_PL_BB_Build,
        #     "Other Metrics": df_Inputs_Other_Metrics,
        #     "Availability Borrower": df_Availability_Borrower,
        #     "Portfolio LeverageBorrowingBase": df_Inputs_Portfolio_LeverageBorrowingBase,
        #     "Obligors' Net Capital": df_Obligors_Net_Capital,
        #     "PL BB Results": pl_bb_results,
        #     "Subscription BB": df_subscription_bb,
        #     "PL_BB_Results_Security": pl_bb_result_security_df,
        #     "Inputs Industries": input_industries_df,
        #     "Advance Rates": advance_rates_df,
        #     "Concentration Limits": concentration_limits_df,
        #     "Principle Obligations": principle_obligation_df,
        #     "Pricing": pricing_df
        # }
        pickled_xl_df_map = pickle.dumps(xl_df_map)

        
        included_excluded_assets = pcofDashboardService.pcof_included_excluded_assets(xl_df_map)

        dt_string = extracted_base_data_info.report_date.strftime("%m_%d_%Y")
        base_data_file = BaseDataFile(
            user_id=1,
            file_data=pickled_xl_df_map,
            fund_type='PCOF',
            file_name ='PCOF Base data ' + dt_string,
            included_excluded_assets_map=included_excluded_assets,
            closing_date=extracted_base_data_info.report_date,
            extracted_base_data_info_id = bdi_id
        )
        db.session.add(base_data_file)
        db.session.commit()
        print(f'PCOF Base data {dt_string}')
        bb_response = pcofBBCalculator.get_bb_calculation(base_data_file=base_data_file, selected_assets=json.loads(included_excluded_assets)['included_assets'], user_id=1)

        bb_response["base_data_file_id"] = base_data_file.id
        wb2.close()
        os.remove(file_name)

        return ServiceResponse.success(message="Successfully processed calculation.", data=bb_response)


    except Exception as e:
        Log.func_error(e)
        return {
            "success": False,
            "message": "Something went wrong while executing borrowing base trigger for PCOF"
        }
       