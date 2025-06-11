from azure.core.exceptions import ResourceExistsError
import os
import mmap
import re
from io import BytesIO
import pandas as pd
from datetime import datetime
import pytz
# import azure.functions as func
from sqlalchemy import text, join, select, func, bindparam
import threading
import numpy as np
from numerize import numerize
from openpyxl import load_workbook
import openpyxl
import pickle
import json
from datetime import datetime
from werkzeug.datastructures import FileStorage
import time

from source.utility.Util import float_to_currency, float_to_numerized
from source.services.aiIntegration.openai import OpenAIClient
from source.services.commons import commonServices
from source.app_configs import azureConfig
from source.utility.ServiceResponse import ServiceResponse
from source.utility.Log import Log
from models import SourceFiles, Users, db, ExtractedBaseDataInfo, PfltBaseData, PfltBaseDataHistory, PcofBaseData, PcofBaseDataHistory, PsslBaseData, PsslBaseDataHistory, BaseDataMapping, PfltSecurityMapping, BaseDataMappingColumnInfo, BaseDataFile, BaseDataOtherInfo, ColumnMetadataMaster, SheetMetadataMaster, FileMetadataMaster, VaeData
from source.services.diServices import helper_functions
from source.services.diServices import base_data_mapping
from source.services.diServices.PCOF import base_data_extractor as pcof_base_data_extractor
from source.services.diServices.PSSL import base_data_extractor as pssl_base_data_extractor
from source.services.diServices.PCOF import BBTrigger as PCOF_BBTrigger
from source.services.diServices.PSSL import BBTrigger as PSSL_BBTrigger
from source.services.diServices import ColumnSheetMap
from source.services.diServices.ColumnSheetMap import ExtractionStatusMaster
from source.services.PFLT.PfltDashboardService import PfltDashboardService
from source.services.diServices.helper_functions import store_sheet_data, check_data_type, check_value_data_type, infer_data_type

pfltDashboardService = PfltDashboardService()

def upload_src_file_to_az_storage(files, report_date, fund_type):
    try:
        if len(files) == 0:
            return ServiceResponse.error(message = "Please select files.", status_code = 400)
        if not fund_type:
            return ServiceResponse.error(message = "Please select Fund.", status_code = 400)
        
        blob_service_client, blob_client = azureConfig.get_az_service_blob_client()
        company_name = "Pennant"
        fund_names = fund_type
        report_date = datetime.strptime(report_date, "%Y-%m-%d").date()
        source_files_list = []
        for file in files:
            print(file.filename)
            blob_name = f"{company_name}/{file.filename}"
            
            # setting SourceFiles object
            file_name = os.path.splitext(file.filename)[0]
            extension = os.path.splitext(file.filename)[1]
            file_bytes = file.read() 
            file.seek(0, 2)  # Move to the end of the file
            file_size = file.tell()  # Get the size in bytes
            file.seek(0)
            company_id = 1
            is_validated = False
            is_extracted = False
            uploaded_by = 1
            file_type = None


            # upload blob in container
            blob_client.upload_blob(name=blob_name, data=file)
            file_url = blob_client.url + '/' + blob_name
            # file_url = '/'
            # add details of files in db
            source_file = SourceFiles(file_name=file_name, extension=extension, report_date=report_date, file_url=file_url, file_size=file_size, company_id=company_id, fund_types=fund_names, is_validated=is_validated, is_extracted=is_extracted, extraction_status='In Progress', uploaded_by=uploaded_by, file_type=file_type)

            source_files_list.append({
                "source_file": source_file,
                "file": file_bytes
            })

            db.session.add(source_file)
            db.session.commit()
            db.session.flush()  
            db.session.refresh(source_file)
            
        return ServiceResponse.success(message = "Files uploaded successfully", data=source_files_list)

    except ResourceExistsError as ree:
        Log.func_error(e=ree)
        return ServiceResponse.error(message="Files with same name already exist.", status_code=409)
    except Exception as e:
        Log.func_error(e=e)
        return ServiceResponse.error(message="Could not upload files.", status_code = 500)
    
def get_blob_list(fund_type):
    company_id = 1 # for Penennt

    # If fund type is not provided -> fetching all records 
    source_files_query = db.session.query(
            SourceFiles.id,
            SourceFiles.file_name,
            SourceFiles.extension,
            SourceFiles.uploaded_at,
            SourceFiles.uploaded_by,
            SourceFiles.fund_types,
            SourceFiles.file_type,
            SourceFiles.report_date,
            SourceFiles.extraction_status,
            Users.display_name,
            SourceFiles.validation_info
        ).join(Users, Users.user_id == SourceFiles.uploaded_by).filter(SourceFiles.is_deleted == False, SourceFiles.company_id == company_id, SourceFiles.is_archived == False)
    
    if fund_type:
        source_files_query = source_files_query.filter(SourceFiles.fund_types.any(fund_type))

    source_files = source_files_query.order_by(SourceFiles.uploaded_at.desc()).all()

    list_table = {
        "columns": [{
            "key": "fund", 
            "label": "Fund",
        }, {
            "key": "file_name", 
            "label": "File Name",
        }, {
            "key": "report_date", 
            "label": "Report Date",
        }, {
            "key": "uploaded_at", 
            "label": "Uploaded Date",
        }, {
            "key": "uploaded_by", 
            "label": "Uploaded by",
        }, {
            "key": "extraction_status",
            "label": "Extraction Status"
        }], 
        "data": []
    }

    for source_file in source_files:
        list_table["data"].append({
            "file_id": source_file.id,
            "file_name": source_file.file_name + source_file.extension, 
            "uploaded_at": source_file.uploaded_at.strftime("%Y-%m-%d"),
            "report_date": source_file.report_date.strftime("%Y-%m-%d"),
            "fund": source_file.fund_types,
            "source_file_type": source_file.file_type,
            "extraction_status": source_file.extraction_status,
            "uploaded_by": source_file.display_name,
            "validation_info": source_file.validation_info
        })
    
    return ServiceResponse.success(data=list_table)

def contains_cash(string):
    return "cash" in string.lower()

def contains_master_comp(string):
    return "master" in string.lower()

def get_data(blob_data, sheet_name, output_file_name, args):
    try:
        # container_client = blob_service_client.get_container_client('on-pepper')
        # blob_client = container_client.get_blob_client(file_path)
        # blob_data = blob_client.download_blob().readall()
        df = pd.read_excel(blob_data, sheet_name=sheet_name)
        for name in args:
            if df.eq(name).any(axis=1).any():
                first_occurrence_index = df.eq(name).any(axis=1).idxmax()
            else:
                first_occurrence_index = None 
            if first_occurrence_index is None:
                continue
            else:
                break
        if first_occurrence_index is None:
            return {"success_status": False, "error": "No such column found", "dataframe": None}
        
        new_df = df.loc[first_occurrence_index + 1:].reset_index(drop=True)
        new_df.columns = df.loc[first_occurrence_index]
        return {"success_status": True, "error": None, "dataframe": new_df}
    except Exception as e:
        return {"success_status": False, "error": str(e), "dataframe": None}

def update_column_names(df, sheet_name, sheet_column_mapper):
    try:
        column_level_map = sheet_column_mapper[sheet_name]
        columns = []
        
        for index, column in enumerate(df.columns):
            column_initials = ""
            for heading in list(column_level_map.keys()):
                if index in range(column_level_map[heading][1], column_level_map[heading][2]):
                # if (index >= heading[1] and index < heading[2]):
                    if column_level_map[heading][0] == -1:
                        column_initials = column_initials + heading + " "
                    else:
                        column_initials = column_initials + "[" + "".join(word[0].upper() for word in heading.split()) + "] "
            column_name = column_initials + str(column)
            columns.append(column_name.strip())

        df.columns = columns
        df = df.round(3)
        return df
    
    except Exception as e:
        Log.func_error(e)



def extract_and_store(file_ids, sheet_column_mapper, extracted_base_data_info, fund_type):
    try:
        engine = db.get_engine()
        fund_name = fund_type
        start_time = datetime.now()

        cash_file_details = None
        master_comp_file_details = None
        market_book_file_details = None
        master_rating_details = None

        for file_id in file_ids:
            file_details = SourceFiles.query.filter_by(id=file_id).first()
            file_type = file_details.file_type
            if (file_type == 'cashfile'):
                cash_file_details = file_details
            elif file_type == 'master_comp':
                master_comp_file_details = file_details
            elif file_type == 'market_book_file':
                market_book_file_details = file_details
            elif file_type == 'master_ratings':
                master_rating_details = file_details
            print(file_type)
           
        
        
        # update security mapping table
        # helper_functions.update_security_mapping(engine)

        if fund_name == "PCOF":
            if master_comp_file_details == None or market_book_file_details == None:
                raise Exception('Proper files not selected.')
            service_response = pcof_base_data_extractor.map_and_store_base_data(engine, extracted_base_data_info, master_comp_file_details, market_book_file_details)
            if service_response["success"]:
                extracted_base_data_info.status = ExtractionStatusMaster.COMPLETED.value
            else:
                raise Exception(service_response.get("message"))
        elif fund_name == 'PSSL':
            if cash_file_details == None or master_comp_file_details == None or master_rating_details == None:
                raise Exception('Proper files not selected.')
            service_response = pssl_base_data_extractor.map_and_store_base_data(engine, extracted_base_data_info, master_comp_file_details, cash_file_details, master_rating_details)
            if service_response["success"]:   
                extracted_base_data_info.status = ExtractionStatusMaster.COMPLETED.value
            else:
                raise Exception(service_response.get("message"))
        else:
            if cash_file_details == None or master_comp_file_details == None or market_book_file_details == None or master_rating_details == None:
                raise Exception('Proper files not selected.')
            base_data_mapping.soi_mapping(engine, extracted_base_data_info, master_comp_file_details, cash_file_details, market_book_file_details, master_rating_details)
            extracted_base_data_info.status = ExtractionStatusMaster.COMPLETED.value
        # else:
            # extracted_base_data_info.status = "repeated"
        
        db.session.add(extracted_base_data_info)
        db.session.commit()
        end_time = datetime.now()
        time_difference = (end_time - start_time).total_seconds() * 10**3
        print('successfully stored base data')
        existing_record = BaseDataOtherInfo.query.filter_by(fund_type=fund_type).order_by(BaseDataOtherInfo.created_at.desc()).first()
        if existing_record:
            determination_date = existing_record.determination_date
            other_data = existing_record.other_info_list
            add_base_data_other_info(
                extracted_base_data_info.id,
                determination_date,
                fund_type, 
                other_data
            )
    except Exception as e:
        extracted_base_data_info.status = "Failed"
        extracted_base_data_info.failure_comments = str(e)
        db.session.add(extracted_base_data_info)
        db.session.commit()
        raise Exception(e)

def get_sheet_data(blob_data, sheet_name, output_file_name, args):
    try:
        df = blob_data
        for name in args:
            if df.eq(name).any(axis=1).any():
                first_occurrence_index = df.eq(name).any(axis=1).idxmax()
            elif name in df.columns:
                first_occurrence_index = 0
            else:
                first_occurrence_index = None 
            if first_occurrence_index is None:
                continue
            else:
                break
        if first_occurrence_index is None:
            return {"success_status": False, "error": "No such column found", "dataframe": None}
        
        new_df = df if first_occurrence_index == 0 else df.loc[first_occurrence_index + 1:].reset_index(drop=True)
        new_df.columns = df.columns if first_occurrence_index == 0 else df.loc[first_occurrence_index]
        return {"success_status": True, "error": None, "dataframe": new_df}
    except Exception as e:
        Log.func_error(e)
        return {"success_status": False, "error": str(e), "dataframe": None}
    
def get_file_type(sheet_name_list):
    cashFileSheetList = {"US Bank Holdings", "Client Holdings"} 
    masterCompSheetList = {"Borrower Stats", "Securities Stats", "SOI Mapping"}
    # marketValueSheetList = {"Sheet1"}
    marketValueSheetList = {"Market and Book Value"}
    masterRatingsList = {"Master Ratings"}

    if any(sheet in sheet_name_list for sheet in masterRatingsList):
        return "master_ratings", list(masterRatingsList)
    elif cashFileSheetList.issubset(set(sheet_name_list)):  
        return "cashfile", list(cashFileSheetList)
    elif masterCompSheetList.issubset(set(sheet_name_list)): 
        return "master_comp", list(masterCompSheetList)
    elif (set(sheet_name_list)).issubset(marketValueSheetList): 
        return "market_book_file", list(marketValueSheetList)

def sheet_data_extract(db_source_file, uploaded_file, updated_column_df, sheet_column_mapper, args):
    try:
        extrcted_df = {}

        uploaded_file.seek(0)
        sheet_df_map = pd.read_excel(uploaded_file, sheet_name=None)

        sheet_name_list = list(sheet_df_map.keys())

        file_type, required_sheets = get_file_type(sheet_name_list)

        for sheet in required_sheets:
            column_level_map = sheet_column_mapper[sheet]
            df_result = get_sheet_data(sheet_df_map[sheet], sheet, column_level_map, args)
            if df_result["success_status"] is True:
                extrcted_df[sheet] = df_result["dataframe"]
                extracted_df = df_result["dataframe"].copy(deep=True)
                updated_col_df = update_column_names(extracted_df, sheet, sheet_column_mapper)
                updated_col_df["source_file_id"] = db_source_file.id
                db_source_file.file_type = file_type
                updated_column_df[sheet] = updated_col_df
            print(sheet + ' extracted')

    except Exception as e:
        Log.func_error(e)
        raise Exception(e)

def extract_source_file(file_value):
    try:
        print("inside", file_value["source_file"])
        sheet_column_mapper = ColumnSheetMap.sheet_column_mapper
        args = ['Company', "Security", "CUSIP", "Asset ID", "SOI Name", "Family Name", "Asset", "Issuer", "MVMinusBV", "Master Comps"]
        updated_column_df = {}

        # for file_value in file_list:
        db_source_file = file_value.get("source_file")
        print("db_source_file", db_source_file)
        uploaded_file_bytes =  file_value.get("file")
        uploaded_file_stream = BytesIO(uploaded_file_bytes)
        uploaded_file = FileStorage(stream=uploaded_file_stream, filename="example.txt", content_type="text/plain")

        sheet_data_extract(db_source_file, uploaded_file, updated_column_df, sheet_column_mapper, args)
        
        return ServiceResponse.success(data=updated_column_df)
    
    except Exception as e:
        print(str(e))
        return ServiceResponse.error()

def extract_base_data(file_ids, fund_type):
    base_data_info_id = None
    try:
    # initialization
        sheet_column_mapper = ColumnSheetMap.sheet_column_mapper
        #--------------------------------
        if len(file_ids) == 0:
            return ServiceResponse.error(message='No files selected.')
        ini_file = SourceFiles.query.filter_by(id = file_ids[0]).first()
        report_date = ini_file.report_date
        company_id = ini_file.company_id
        extracted_base_data_info = ExtractedBaseDataInfo(report_date=report_date, fund_type=fund_type, status=ExtractionStatusMaster.IN_PROGRESS.value, company_id = 1, files = file_ids)
        db.session.add(extracted_base_data_info)
        db.session.commit()
        db.session.refresh(extracted_base_data_info)
        base_data_info_id = extracted_base_data_info.id

        extract_and_store(file_ids=file_ids, sheet_column_mapper=sheet_column_mapper, extracted_base_data_info=extracted_base_data_info, fund_type=fund_type)

        # threading.Thread(target=extract_and_store, kwargs={
        #     'file_ids': file_ids,
        #     'sheet_column_mapper': sheet_column_mapper,
        #     'extracted_base_data_info': extracted_base_data_info,
        #     'fund_type': fund_type}
        # ).start()

        # extract_and_store(file_ids = file_ids, sheet_column_mapper = sheet_column_mapper, extracted_base_data_info = extracted_base_data_info)

        response_data = {
            "id": extracted_base_data_info.id,
            "report_date": report_date.strftime("%Y-%m-%d"),
            "company_id": company_id
        }
        return ServiceResponse.success(message="Base Data extracted. Redirecting...", data=response_data)
    except Exception as e:
        if base_data_info_id:
            extracted_base_data_info = ExtractedBaseDataInfo.query.filter_by(id = base_data_info_id).first()
            extracted_base_data_info.status = ExtractionStatusMaster.FAILED.value
            extracted_base_data_info.failure_comments = str(e)
            db.session.add(extracted_base_data_info)
            db.session.commit()
        return ServiceResponse.error(message=str(e))

def persist_old_base_data(fund_type):
    try:
        # 1. get last base data and previous base data
        extracted_base_data_info = ExtractedBaseDataInfo.query.filter_by(fund_type = fund_type, status = 'Completed').order_by(ExtractedBaseDataInfo.id.desc()).limit(2).all()
        # if no previous base data then return
        if len(extracted_base_data_info) < 2:
            return ServiceResponse.error(message='No older base data present.')
        base_data_table = None
        unique_identifier = []
        match fund_type:
            case "PFLT":
                base_data_table = PfltBaseData
                unique_identifier = ['obligor_name', 'security_name', 'loan_type']
            case "PCOF":
                base_data_table = PcofBaseData
                unique_identifier = ['investment_name', 'issuer']
            case "PSSL":
                base_data_table = PsslBaseData
                unique_identifier = ['borrower', 'loan_type']
        current_base_data = base_data_table.query.filter(base_data_table.base_data_info_id == extracted_base_data_info[0].id).order_by(base_data_table.id).all()
        previous_base_data = base_data_table.query.filter(base_data_table.base_data_info_id == extracted_base_data_info[1].id).order_by(base_data_table.id).all()
        column_keys = [column.name for column in base_data_table.__table__.columns]
        for data in current_base_data:
            for prev_data in previous_base_data:
                data_prev_record = None
                for column in unique_identifier:
                    # if getattr(data, column) is None or getattr(prev_data, column) is None:
                    #     data_prev_record = None
                    #     break
                    if getattr(data, column) != getattr(prev_data, column):
                        data_prev_record = None
                        break
                    else:
                        data_prev_record = prev_data
                if data_prev_record is not None:
                    break
            if data_prev_record is not None:
                for column in column_keys:
                    value = getattr(data, column)
                    temp = value
                    if value is None:
                        if getattr(data_prev_record, column) is not None:
                            temp = getattr(data_prev_record, column)
                    setattr(data, column, temp)
        db.session.commit()
        return ServiceResponse.success(message="Base data updated")
    except Exception as e:
        raise Exception(e)



def get_base_data(info_id):
    try:
        # datetime_obj = datetime.strptime(report_date, "%Y-%m-%d")
       
        base_data_info = ExtractedBaseDataInfo.query.filter_by(id = info_id).first()
        if base_data_info.fund_type == 'PFLT':
            sheet_name = "Loan List"
        elif base_data_info.fund_type == 'PCOF':
            sheet_name = "PL BB Build"
        else:
            sheet_name = "Portfolio"
 
        base_data_mapping = db.session.query(
            BaseDataMapping.bdm_id,
            BaseDataMapping.bd_column_lookup,
            BaseDataMapping.bd_column_name,
            BaseDataMapping.bd_column_datatype,
            BaseDataMapping.bd_column_unit,
            BaseDataMapping.bd_column_is_required,
            BaseDataMapping.is_one_time_input,
            BaseDataMapping.is_on_going_input_rarely_updated,
            BaseDataMapping.is_on_going_input,
            BaseDataMapping.description,
            BaseDataMapping.is_editable,
            BaseDataMapping.sf_sheet_name,
            BaseDataMapping.sf_column_name,
            BaseDataMappingColumnInfo.sequence,
            BaseDataMappingColumnInfo.is_selected
        ).join(BaseDataMapping, BaseDataMapping.bdm_id == BaseDataMappingColumnInfo.bdm_id).filter(BaseDataMapping.fund_type == base_data_info.fund_type, BaseDataMapping.bd_sheet_name == sheet_name).order_by(BaseDataMappingColumnInfo.sequence).all()
 
 
        card_data = []
 
        engine = db.get_engine()
       
        if base_data_info.fund_type == 'PFLT':
            base_data = PfltBaseData.query.filter(PfltBaseData.base_data_info_id == info_id).order_by(PfltBaseData.id).all()
            HistoryData = PfltBaseDataHistory
            with engine.connect() as connection:
                result = connection.execute(text('''
                    select count(distinct pbd.obligor_name) as no_of_obligors,
                            count(distinct pbd.security_name) as no_of_assets,
                            sum(pbd.total_commitment::float) as total_commitment,
                            sum(pbd.outstanding_principal::float) as total_outstanding_balance,
                            (select count(distinct ssubh."Security/Facility Name") from source_files sf left join sf_sheet_us_bank_holdings ssubh on ssubh.source_file_id = sf.id left join pflt_security_mapping psm on psm.cashfile_security_name = ssubh."Security/Facility Name" where sf.id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :info_id) and psm.id is null and file_type = 'cashfile') as unmapped_records
                    from pflt_base_data pbd
                    where pbd.base_data_info_id = :info_id'''), {'info_id': info_id}).fetchall()
                final_result = result[0]
            no_of_obligors, no_of_assets, total_commitment, total_outstanding_balance, unmapped_records = final_result
            card_data = [{
                "No of Obligors": no_of_obligors,
                "No of Securities": no_of_assets,
                "Total Commitment": '$' + numerize.numerize(total_commitment, 2) if total_commitment is not None else 0,
                "Total Outstanding Balance": '$' + numerize.numerize(total_outstanding_balance, 2)  if total_outstanding_balance is not None else 0,
                "Unmapped Securities": unmapped_records,
                "Report Date": base_data_info.report_date.strftime("%Y-%m-%d"),
                "Fund Type": base_data_info.fund_type
            }]
        if base_data_info.fund_type == 'PSSL':
            base_data = PsslBaseData.query.filter(PsslBaseData.base_data_info_id == info_id).order_by(PsslBaseData.id).all()
            HistoryData = PsslBaseDataHistory
            card_data = [{
                "Report Date": base_data_info.report_date.strftime("%Y-%m-%d"),
                "Fund Type": base_data_info.fund_type
            }]
        if base_data_info.fund_type == 'PCOF':
            base_data = PcofBaseData.query.filter_by(base_data_info_id = info_id).order_by(PcofBaseData.id).all()
            HistoryData = PcofBaseDataHistory
            with engine.connect() as connection:
                result = connection.execute(text('''
                    select count(distinct pbd.issuer) as no_of_issuers,
                            count(distinct pbd.investment_name) as no_of_investments,
                            (select count(distinct pbd2.issuer) as eligible_issuers from pcof_base_data pbd2 where pbd2.base_data_info_id = :info_id and pbd2.is_eligible_issuer = 'Yes') as eligible_issuers,
                            sum(NULLIF(pbd.investment_cost, '')::float) as total_investment_cost,
                            sum(NULLIF(pbd.investment_par, '')::float) as total_investment_par,
                            sum(NULLIF(pbd.investment_external_valuation, '')::float) as total_investment_external_valuation,
                            sum(NULLIF(pbd.investment_internal_valuation, '')::float) as total_investment_internal_valuation
                            --sum(pbd.total_commitment::float) as total_commitment,
                            --sum(pbd.outstanding_principal::float) as total_outstanding_balance,
                            --(select count(distinct ssm.Issuer_Name) from source_files sf left join sf_sheet_marketbook_1 ssm on ssubh.source_file_id = sf.id left join pflt_security_mapping psm on psm.cashfile_security_name = ssubh."Security/Facility Name" where sf.id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :info_id) and psm.id is null and file_type = 'cashfile') as unmapped_records
                    from pcof_base_data pbd
                    where pbd.base_data_info_id = :info_id'''), {'info_id': info_id}).fetchall()
                final_result = result[0]
            no_of_issuers, no_of_investments, eligible_issuers, total_investment_cost, total_investment_par, total_investment_external_valuation, total_investment_internal_valuation = final_result
            card_data = [{
                "No of Issuers": no_of_issuers,
                "No of Investments": no_of_investments,
                "Eligible Issuers": eligible_issuers,
                "Total Cost": '$' + float_to_numerized(total_investment_cost),
                "Total Par": '$' + float_to_numerized(total_investment_par),
                "Total External Valuation": '$' + float_to_numerized(total_investment_external_valuation),
                "Total Internal Valuation": '$' + float_to_numerized(total_investment_internal_valuation),
                "Report Date": base_data_info.report_date.strftime("%Y-%m-%d"),
                # "Unmapped Securities": unmapped_records,
                "Fund Type": base_data_info.fund_type
            }]
 
        temp = []
        # print(base_data[0])
        for b in base_data:
            t = b.__dict__
            del t['_sa_instance_state']
            old_data = HistoryData.query.filter_by(id = b.id).order_by(HistoryData.done_at.asc()).limit(1).first()
            t1 = None
            if old_data:
                t1 = old_data.__dict__
            for key in t:
                val = t[key]
                old_value = val
                if t1 and t1[key] != t[key]:
                    old_value = t1[key]
                numerized_val = None
                if isinstance(val, (int, float, complex)) and not isinstance(val, bool):
                    numerized_val = numerize.numerize(val, 2)
                    # t[key] = val
                elif isinstance(val, str):
                    if val.replace(".", "").replace("-", "").isnumeric():
                        try:
                            numerized_val = numerize.numerize(float(val), 2)
                        except Exception as e:
                            print(e, val, type(val))
                        # t[key] = val
                col_details = next((b for b in base_data_mapping if b.bd_column_lookup == key), None)
                if col_details is not None and col_details[4] == 'currency':
                    numerized_val = '$' + numerized_val if numerized_val is not None else val
                t[key] = {
                    "meta_info": True,
                    "value": val,
                    "display_value": numerized_val if numerized_val else val,
                    "title": val,
                    "old_value": old_value
                }
            # t['report_date'] = t['report_date'].strftime("%Y-%m-%d")
            # t['created_at'] = t['created_at'].strftime("%Y-%m-%d")
            temp.append(t)
        base_data_table = {
            "columns": [{
                "key": column.bd_column_lookup,
                "label": column.bd_column_name,
                "datatype": column.bd_column_datatype,
                "unit": column.bd_column_unit,
                "isEditable": column.is_editable,
                "bdm_id": column.bdm_id,
                "is_selected": column.is_selected,
                "is_one_time_input": column.is_one_time_input,
                "is_on_going_input_rarely_updated": column.is_on_going_input_rarely_updated,
                "is_on_going_input": column.is_on_going_input,
                "description": column.description,
                "bd_column_is_required": column.bd_column_is_required,
                "sf_sheet_name": column.sf_sheet_name,
                "sf_column_name": column.sf_column_name
            } for column in base_data_mapping],
            "data": temp
        }
 
        # print(type(base_data_table))
       
       
        # for index, row in base_data.iterrows():
        #     row_data = {}
           
        #     for col, value in row.items():
        #         if value != value:
        #             value = ""
        #         if isinstance(value, (int, float)):
        #                 row_data[col.replace(" ", "_")] = value
        #         elif isinstance(value, datetime):
        #             row_data[col.replace(" ", "_")] = value.strftime("%Y-%m-%d")
        #         else:
        #             row_data[col.replace(" ", "_")] = value
        #     base_data_table["data"].append(row_data)
        result = {
            "base_data_table": base_data_table,
            "report_date": base_data_info.report_date.strftime("%Y-%m-%d"),
            "fund_type": base_data_info.fund_type,
            "card_data": card_data
        }
        return ServiceResponse.success(data=result, message="Base Data")
    except Exception as e:
        raise Exception(e)
    
def change_bd_col_seq(updated_sequence):
    try:
        updated_sequence_list = []
        for change in updated_sequence:
            bdm_id = change.get("bdm_id")
            sequence = change.get("sequence")
            base_data_mapping_info = BaseDataMappingColumnInfo.query.filter_by(bdm_id = bdm_id).first()
            base_data_mapping_info.sequence = sequence
            updated_sequence_list.append(base_data_mapping_info)
        
        db.session.add_all(updated_sequence_list)
        db.session.commit()

        return ServiceResponse.success(message = "Base Data Column Sequence Changed Successfully")
    except Exception as e:
        raise Exception(e)
    
def get_base_data_col(fund_type):
    try:
        base_data_columns_data = db.session.query(
            BaseDataMapping.bdm_id,
            BaseDataMapping.bd_column_lookup,
            BaseDataMapping.bd_column_name,
            BaseDataMapping.is_editable,
            BaseDataMappingColumnInfo.sequence,
            BaseDataMappingColumnInfo.is_selected
        ).join(BaseDataMapping, BaseDataMapping.bdm_id == BaseDataMappingColumnInfo.bdm_id).filter(BaseDataMappingColumnInfo.fund_type == fund_type).order_by(BaseDataMappingColumnInfo.sequence).all()
        
        base_data_columns = [{
            "key": column.bd_column_lookup,
            "label": column.bd_column_name,
            "isEditable": column.is_editable,
            "bdm_id": column.bdm_id,
            "is_selected": column.is_selected
        } for column in base_data_columns_data]

        return ServiceResponse.success(data=base_data_columns, message="Base Data Columns")
    except Exception as e:
        Log.func_error(e)
        return ServiceResponse.error(message="Could not get base data columns")
    
def update_bd_col_select(selected_col_ids, fund_type):
    try:
        # making is_selected as false to all records
        BaseDataMappingColumnInfo.query.filter_by(fund_type=fund_type).update({"is_selected": False})

        default_col_seq = 1
        # modified_by = Users.query.filter_by(id=1).first()
        modified_by = 1
            
        # updating is_selected of selected columns to True with default sequence
        selected_bd_col_info_list = []
        for col_id in selected_col_ids:
            selected_bd_col_info = BaseDataMappingColumnInfo.query.filter_by(bdm_id=col_id).first()
            selected_bd_col_info.sequence = default_col_seq
            default_col_seq = default_col_seq + 1
            selected_bd_col_info.is_selected = True
            selected_bd_col_info.modified_at = datetime.now()
            selected_bd_col_info.modified_by = modified_by
            selected_bd_col_info_list.append(selected_bd_col_info)
        db.session.add_all(selected_bd_col_info_list)
        db.session.commit()

        # updating sequence of unselected columns
        unselected_bd_col_info_list = []
        unselected_bd_cols = BaseDataMappingColumnInfo.query.filter_by(is_selected=False, fund_type=fund_type).all()
        for unselected_bd_col in unselected_bd_cols:
            unselected_bd_col.sequence = default_col_seq
            default_col_seq = default_col_seq + 1
            unselected_bd_col_info_list.append(unselected_bd_col)

        db.session.add_all(unselected_bd_col_info_list)
        db.session.commit()

        return ServiceResponse.success(message = "Base Data Column Selection Updated Successfully")

    except Exception as e:
        Log.func_error(e=e)
        db.session.rollback()
        return ServiceResponse.error(message = "Could not Update base data column selection")

def edit_base_data(changes):
    if not changes:
        return ServiceResponse.error(message = "Please provide changes.", status_code = 400)
    for change in changes:
        id = change.get("id")
        fund = change.get("fundType")
        table = None
        match(fund):
            case 'PFLT':
                table = PfltBaseData
            case 'PCOF':
                table = PcofBaseData
            case 'PSSL':
                table = PsslBaseData
            case _:
                return
        if table is None:
            return ServiceResponse.error(message = "Fund not specified", status_code = 400)
        for key in change.keys():
            if key != "id" and key != "fundType":
                value = change.get(key)
                # column = key.replace('_', " ")
                base_data = table.query.filter_by(id=id).first()
                setattr(base_data, key, value)
                db.session.add(base_data)
                db.session.commit()
    return ServiceResponse.success(message = "Base data edited updated successfully")

def get_base_data_mapping(info_id):
    try:
        engine = db.get_engine()
        with engine.connect() as connection:
            base_data_map = pd.DataFrame(connection.execute(text('''select * from pflt_base_data_mapping''')).fetchall())
            df_dict = base_data_map.to_dict(orient='records')
        return ServiceResponse.success(data=df_dict, message="Base Data Map")
    except Exception as e:
        raise Exception(e)

def get_extracted_base_data_info(company_id, extracted_base_data_info_id, fund_type):
    if extracted_base_data_info_id:
        extracted_base_datas = ExtractedBaseDataInfo.query.filter_by(id = extracted_base_data_info_id)
    else:
        extracted_base_datas = ExtractedBaseDataInfo.query.filter_by(company_id = company_id).order_by(ExtractedBaseDataInfo.id.desc())
    if fund_type:
        extracted_base_datas = extracted_base_datas.filter_by(fund_type=fund_type).order_by(ExtractedBaseDataInfo.id.desc())
        
    extracted_base_datas = extracted_base_datas.order_by(ExtractedBaseDataInfo.report_date.desc()).all()
    
    extraction_result = {
        "columns": [{
            "key": "id",
            "label": "#"
        }, {
            "key": "report_date",
            "label": "Report Date"
        }, {
            "key": "fund",
            "label": "Fund"
        }, {
            "key": "extraction_status",
            "label": "Extraction Status"
        }, {
            "key": "extraction_date",
            "label": "Extraction Date"
        }, {
            "key": "source_files",
            "label": "Source Files"
        }],
        "data": []
    }

    for extracted_base_data in extracted_base_datas:
        source_files = SourceFiles.query.filter(SourceFiles.id.in_(extracted_base_data.files)).all()
        files = ''
        file_details = []
        for file in source_files:
            if files == '':
                files = file.file_name
            else:
                files = files + '; ' + file.file_name
            file_details.append({'file_id': file.id, 'file_name': file.file_name, 'file_type': file.file_type, 'extension': file.extension})
        extraction_result["data"].append({
            "id": extracted_base_data.id,
            "report_date": extracted_base_data.report_date.strftime("%Y-%m-%d"),
            "fund": extracted_base_data.fund_type,
            "extraction_status": extracted_base_data.status,
            "comments": extracted_base_data.failure_comments,
            "extraction_date": extracted_base_data.extraction_date.strftime("%Y-%m-%d"),
            "source_files": files,
            "source_file_details":  file_details
        })
    
    return ServiceResponse.success(data=extraction_result)

def get_pflt_sec_mapping():
    engine = db.get_engine()
    with engine.connect() as connection:
        pflt_sec_mapping_df = pd.DataFrame(connection.execute(text('select id, soi_name, master_comp_security_name, family_name, security_type, cashfile_security_name from pflt_security_mapping where company_id = 1 order by soi_name ASC')).fetchall())

        unmapped_securities = pd.DataFrame(connection.execute(text('''select distinct "Security/Facility Name" as cashfile_securities from sf_sheet_us_bank_holdings pubh left join pflt_security_mapping psm on psm.cashfile_security_name = pubh."Security/Facility Name" where psm.id is null''')).fetchall())
    
    columns_data = [
        {
            'key': "soi_name",
            'label': "SOI Name",
            'isEditable': False,
            'isRequired': True
        }, {
            'key': "master_comp_security_name",
            'label': "Master Comp Security Name",
            'isEditable': False
        },  {
            'key': "family_name",
            'label': "Family Name",
            'isEditable': True
        }, {
            'key': "security_type",
            'label': "Security Type",
            'isEditable': False
        }, {
            'key': "cashfile_security_name",
            'label': "Cash File Security Name",
            'isEditable': True
        }
    ]
    pflt_sec_mapping_table = {
        "columns": columns_data,
        "data": []
    }
    pflt_sec_mapping_df.columns = pflt_sec_mapping_df.columns.str.replace(" ", "_")
    pflt_sec_mapping_df = pflt_sec_mapping_df.replace({np.nan: None})
    df_dict = pflt_sec_mapping_df.to_dict(orient='records')
    pflt_sec_mapping_table["data"] = df_dict
    pflt_sec_mapping_table["unmapped_securities"] = unmapped_securities.to_dict(orient='records')

    return ServiceResponse.success(data=pflt_sec_mapping_table, message="pflt security mapping")

def edit_pflt_sec_mapping(changes):
    engine = db.get_engine()
    modified_by = 1 # currently hard coding this variable
    for change in changes:
        id = change.get("id")
        for key in change.keys():
            if key != "id":
                value = change.get(key)
                with engine.connect() as connection:
                    connection.execute(text(f"UPDATE pflt_security_mapping SET {key} = :value, modified_by = :modified_by, modified_at = now() WHERE id = :id"), {"value": value, "id":id, 'modified_by': modified_by})
                    connection.commit()
    
    return ServiceResponse.success(message="PFLT security mapping edited successfully")

def add_pflt_sec_mapping(cashfile_security_name, family_name, master_comp_security_name, security_type, soi_name):
    company_id = 1
    modified_by = 1
    timestamp = datetime.now(pytz.UTC)

    pfltSecurityMapping = PfltSecurityMapping(company_id=company_id, soi_name=soi_name, master_comp_security_name=master_comp_security_name, family_name=family_name, security_type=security_type, cashfile_security_name=cashfile_security_name, modified_by=modified_by, modified_at=timestamp)

    db.session.add(pfltSecurityMapping)
    db.session.commit()
    
    return ServiceResponse.success(message="PFLT security mapping added successfully")

def get_source_file_data(file_id, file_type, sheet_name):
    # ["US Bank Holdings", "Client Holdings"] for Cash
    # ["Borrower Stats", "Securities Stats", "PFLT Borrowing Base"] for Master Comp
    if file_type == "cashfile" and sheet_name == None:
        sheet_name = "US Bank Holdings"
    if file_type == "master_comp" and sheet_name == None:
        sheet_name = "Borrower Stats"
    print(sheet_name)
    match sheet_name:
        case "US Bank Holdings":
            table_name = "sf_sheet_us_bank_holdings"
        case "Client Holdings":
            table_name = "sf_sheet_client_holdings"
        case "Borrower Stats":
            table_name = "sf_sheet_borrower_stats"
        case "Securities Stats":
            table_name = "sf_sheet_securities_stats"
        case "PFLT Borrowing Base":
            table_name = "sf_sheet_pflt_borrowing_base"

    source_file_table_data = {}

    engine = db.get_engine()
    with engine.connect() as connection:
        sheet_df = pd.DataFrame(connection.execute(text(f'select * from "{table_name}" where source_file_id = {file_id}')).fetchall())

    colummns = [
        {
            'key': column.replace(' ', '_'), 
            'label': column
        } for column in sheet_df.columns]
    source_file_table_data['columns'] = colummns

    sheet_df.columns = sheet_df.columns.str.replace(" ", "_")
    sheet_df = sheet_df.replace({np.nan: None})
    df_dict = sheet_df.to_dict(orient='records')
    source_file_table_data['data'] = df_dict
    
    return ServiceResponse.success(data=source_file_table_data)


def get_source_file_data_detail(ebd_id, column_key, data_id):
    try:
        extract_base_data_info = ExtractedBaseDataInfo.query.filter_by(id = ebd_id).first()
        fund_type = extract_base_data_info.fund_type
        obligor_name = None
        security_name = None
        additional_column_condition = ''
        match fund_type:
            case "PCOF":
                table_name = "pcof_base_data"
                history_table_name = "pcof_base_data_history"
            case "PFLT":
                table_name = "pflt_base_data"
                history_table_name = "pflt_base_data_history"
            case "PSSL":
                table_name = "pssl_base_data"
                history_table_name = "pssl_base_data_history"
        engine = db.get_engine()
        with engine.connect() as connection:
            df = pd.DataFrame(connection.execute(text(f'select sf.id, sf.file_type, sf.file_name, sf."extension", pbdm.bd_column_name, pbdm.bd_column_lookup, pbdm.sf_sheet_name, pbdm.sf_column_name, pbdm.sd_ref_table_name, case when pbdm.sf_column_lookup is null then pbdm.sf_column_name else pbdm.sf_column_lookup end as sf_column_lookup, sf_column_categories, formula, ebdi.files from extracted_base_data_info ebdi join source_files sf on sf.id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id) join base_data_mapping pbdm on pbdm.sf_file_type = sf.file_type where ebdi.id = :ebd_id and pbdm.bd_column_lookup = :column_key and pbdm.fund_type = ebdi.fund_type'), {'ebd_id': ebd_id, 'column_key': column_key}).fetchall())
            bd_df = pd.DataFrame(connection.execute(text(f'select * from {table_name} where id = :data_id'), {'data_id': data_id}).fetchall())
        condition_values = None
        match fund_type:
            case "PCOF":
                additional_column_condition = 'and ss."Security" = :security_name'
                condition_values = {'obligor_name': bd_df["issuer"][0], 'security_name': bd_df["investment_name"][0], 'ebd_id': ebd_id}
            case "PFLT":
                additional_column_condition = 'and ss."Security" = :security_name'
                condition_values = {'obligor_name': bd_df["obligor_name"][0], 'security_name': bd_df["security_name"][0], 'ebd_id': ebd_id}
            case "PSSL":
                condition_values = {'obligor_name': bd_df["borrower"][0], 'ebd_id': ebd_id}
        df = df.replace({np.nan: None})
        df_dict = df.to_dict(orient='records')
        try:
            table_name = df_dict[0]['sd_ref_table_name']
            sd_col_name = df_dict[0]['sf_column_name']
            identifier_col_name = None
            if table_name == 'sf_sheet_client_holdings':
                identifier_col_name = 'ch."Issuer/Borrower Name"'
                alias = "ch"
            elif table_name == 'sf_sheet_us_bank_holdings':
                identifier_col_name = 'usbh."Issuer/Borrower Name"'
                alias = "usbh"
            elif table_name == 'sf_sheet_securities_stats':
                identifier_col_name = 'ss."Security"'
                alias = "ss"
            elif table_name == 'sf_sheet_borrower_stats':
                identifier_col_name = 'bs."Company"'
                alias = "bs"
            elif table_name == 'sf_sheet_pflt_borrowing_base':
                identifier_col_name = 'pbb."Security"'
                alias = "pbb"
            elif table_name == 'sf_sheet_marketbook_1':
                identifier_col_name = 'ssm."Issuer_Name"'
                alias = "ssm"
            sd_df_dict = None
            if identifier_col_name is not None:
                if df_dict[0]['sf_column_categories'] is not None and len(df_dict[0]['sf_column_categories']) > 0:
                    sd_col_name = df_dict[0]['sf_column_categories'][0] + " " + sd_col_name
                sd_col_name = alias + "." + '"' + sd_col_name + '"'
                with engine.connect() as connection:
                    sd_df = pd.DataFrame(connection.execute(text(f'''WITH usbh_filtered AS (SELECT * FROM sf_sheet_us_bank_holdings WHERE source_file_id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id)),
                    ch_filtered AS (SELECT * FROM sf_sheet_client_holdings WHERE source_file_id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id)),
                    ss_filtered AS (SELECT * FROM sf_sheet_securities_stats WHERE source_file_id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id)),
                    pbb_filtered AS (SELECT * FROM sf_sheet_pflt_borrowing_base WHERE source_file_id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id)),
                    bs_filtered AS (SELECT * FROM sf_sheet_borrower_stats WHERE source_file_id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id))
                    select distinct {identifier_col_name}, usbh."LoanX ID", {sd_col_name} from usbh_filtered usbh
                    left join ch_filtered ch on ch."Issuer/Borrower Name" = usbh."Issuer/Borrower Name" and ch."Current Par Amount (Issue Currency) - Settled" = usbh."Current Par Amount (Issue Currency) - Settled"
                    left join pflt_security_mapping sm on sm.cashfile_security_name = usbh."Security/Facility Name"
                    left join ss_filtered ss on ss."Security" = sm.master_comp_security_name
                    left join pbb_filtered pbb on pbb."Security" = ss."Security"
                    left join bs_filtered bs on bs."Company" = ss."Family Name"
                    left join (select loan_mapping_cashfile.loan_type, loan_master_cashfile.loan_type as master_loan_type from loan_type_mapping loan_mapping_cashfile
		                left join loan_type_master loan_master_cashfile on loan_master_cashfile.id = loan_mapping_cashfile.master_loan_type_id
		                where (loan_mapping_cashfile.is_deleted = false or loan_mapping_cashfile.is_deleted is null) and loan_master_cashfile.fund_type = '{fund_type}')
                    as t2 on t2.loan_type = ch."Issue Name"
                    left join (select *, loan_master_marketbook.loan_type as master_loan_type from sf_sheet_marketbook_1 ssm 
			            left join loan_type_mapping loan_mapping_marketbook on loan_mapping_marketbook.loan_type = ssm."Asset_Name" and (loan_mapping_marketbook.is_deleted = false or loan_mapping_marketbook.is_deleted is null)
			            left join loan_type_master loan_master_marketbook on loan_master_marketbook.id = loan_mapping_marketbook.master_loan_type_id and loan_master_marketbook.fund_type = '{fund_type}'
			            where ssm.source_file_id in (select unnest(files) from extracted_base_data_info ebdi where ebdi.id = :ebd_id)) as ssm on lower(regexp_replace(ch."Issuer/Borrower Name", '[^a-zA-Z0-9]','', 'g')) = lower(regexp_replace(ssm."Issuer_Name", '[^a-zA-Z0-9]','', 'g')) and t2.master_loan_type = ssm.master_loan_type
                    where usbh."Issuer/Borrower Name" = :obligor_name {additional_column_condition}'''), condition_values).fetchall())
                sd_df_dict = sd_df.to_dict(orient='records')
        except Exception as e:
            print(str(e)[:150])
            Log.func_error(e=e)
        if len(bd_df.columns) > 0 and bd_df['is_manually_added'][0]:
            result = {
                'mapping_data': df_dict[0],
                'source_data': sd_df_dict,
                'is_manual': True
            }
        else:
            result = {
                'mapping_data': df_dict[0],
                'source_data': sd_df_dict,
                'is_manual': False
            }
        if len(df_dict) == 0:
            return ServiceResponse.error(message='No data found.', status_code=404)
        if len(df_dict) > 1:
            return ServiceResponse.error(message='Multiple records found.', status_code=409)
        return ServiceResponse.success(data=result)
    except Exception as e:
        print(str(e)[:150])
        raise Exception(e)

def trigger_bb_calculation(bdi_id):
    try:
        extracted_base_data_info = ExtractedBaseDataInfo.query.filter_by(id=bdi_id).first()
        if extracted_base_data_info.fund_type == 'PCOF':
            pcof_bb_trigger_response = PCOF_BBTrigger.trigger_pcof_bb(bdi_id)
            return pcof_bb_trigger_response
        if extracted_base_data_info.fund_type == 'PSSL':
            pssl_bb_trigger_response = PSSL_BBTrigger.trigger_pssl_bb(bdi_id)
            return pssl_bb_trigger_response
        engine = db.get_engine()
        with engine.connect() as connection:
            df = pd.DataFrame(connection.execute(text(f'select * from pflt_base_data where base_data_info_id = :ebd_id'), {'ebd_id': bdi_id}).fetchall())
            df2 = pd.DataFrame(connection.execute(text(f'select bd_column_name, bd_column_lookup from base_data_mapping where bd_column_lookup is not null and fund_type = \'PFLT\'')).fetchall())
            # haircut_config = pd.DataFrame(connection.execute(text(f'select haircut_level, obligor_tier, "position", value from pflt_haircut_config')).fetchall())
            industry_list = pd.DataFrame(connection.execute(text(f'select industry_no as "Industry No", industry_name as "Industry" from pflt_industry_list')).fetchall())
        df = df.replace({np.nan: None})
        report_date = df['report_date'][0].strftime("%Y-%m-%d")
        df = df.drop('report_date', axis=1)
        df = df.drop('created_at', axis=1)
        df = df.drop('base_data_info_id', axis=1)
        df = df.drop('id', axis=1)
        df = df.drop('company_id', axis=1)
        df = df.drop('created_by', axis=1)
        df = df.drop('modified_by', axis=1)
        df = df.drop('modified_at', axis=1)
        # df['report_date'] = df['report_date'].astype(str)
        # df['created_at'] = df['created_at'].astype(str)
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Loan List"
        file_name = "PFLT Base data - " + report_date + ".xlsx"
        wb.save(file_name)

        path1 = 'PFLT Base data.xlsx'
        path2 = file_name
        wb1 = openpyxl.load_workbook(filename=path1)
        for index, sheet in enumerate(wb1.worksheets):
            ws1 = wb1.worksheets[index]

            wb2 = openpyxl.load_workbook(filename=path2)
            ws2 = wb2.create_sheet(ws1.title)
            for row in ws1:
                for cell in row:
                    ws2[cell.coordinate].value = cell.value
            wb2.save(path2)
        
        rename_df_col = {}
        for index, row in df2.iterrows():
            rename_df_col[row['bd_column_lookup']] = row['bd_column_name']
        df.rename(columns=rename_df_col, inplace=True)

        df["Spread incl. PIK and PIK'able"] = pd.to_numeric(df["Spread incl. PIK and PIK'able"], errors='coerce')
        df["PIK / PIK'able For Floating Rate Loans"] = pd.to_numeric(df["PIK / PIK'able For Floating Rate Loans"], errors='coerce')

        # print(df.dtypes)
        # for c in df.columns:
            # print(c, df[c].dtype)

        # df["Initial TTM EBITDA"] = df["Initial TTM EBITDA"].astype(float)
        # df["Current TTM EBITDA"] = df["Current TTM EBITDA"].astype(float)
        # df["Current Fixed Charge Coverage Ratio"] = df["Current Fixed Charge Coverage Ratio"].astype(float)
        # df["Initial Fixed Charge Coverage Ratio"] = df["Current Fixed Charge Coverage Ratio"].astype(float)

        # df["Spread incl. PIK and PIK'able"].fillna(0, inplace=True)
        # df["PIK / PIK'able For Floating Rate Loans"].fillna(0, inplace=True)
        # df["PIK / PIK'able For Fixed Rate Loans"].fillna(0, inplace=True)
        # df["Base Rate"].fillna(0, inplace=True)

        xl_df_map = {}
        xl_df_map['Loan List'] = df

        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine="openpyxl")
        writer.book = book
        df.to_excel(writer, sheet_name="Loan List", index=False, header=True)
        writer.save()

        base_data_other_info = BaseDataOtherInfo.query.filter_by(extraction_info_id=bdi_id).first()

        inputs_sheet_data = [{
                "INPUTS": "Determination Date",
                "Values": datetime.strptime(base_data_other_info.determination_date[:-5], "%Y-%m-%dT%H:%M:%S")
            }, {
                "INPUTS": "Minimum Equity Amount Floor",
                "Values": float(base_data_other_info.other_info_list["input"].get('minimum_equity_amount_floor', 0))
            }
        ]
        inputs_df = pd.DataFrame(inputs_sheet_data)
        inputs_df.to_excel(writer, sheet_name="Inputs", index=False, header=True)
        writer.save()
        xl_df_map['Inputs'] = inputs_df

        # book = load_workbook(file_name)
        # writer = pd.ExcelWriter(file_name, engine="openpyxl")
        # writer.book = book
        # data.to_excel(writer, sheet_name="Inputs", index=False, header=True)
        # writer.save()

        
        # data = {'Currency': ['USD', 'CAD', 'AUD', 'EUR'], 'Exchange Rate': [1.000000, 0.695230, 0.618820, 1.035400]}
        # data = pd.DataFrame.from_dict(data)
        exchange_rates_data = [
            {
                'Currency': input_data.get('currency'), 
                'Exchange Rate': input_data.get('exchange_rates')
            }
            for input_data in base_data_other_info.other_info_list.get('other_sheet', [])
            if input_data
        ]
        exchange_rates_df = pd.DataFrame(exchange_rates_data)
        exchange_rates_df.to_excel(writer, sheet_name="Exchange Rates", index=False, header=True)
        writer.save()
        xl_df_map['Exchange Rates'] = exchange_rates_df

        # book = load_workbook(file_name)
        # writer = pd.ExcelWriter(file_name, engine="openpyxl")
        # writer.book = book
        # data.to_excel(writer, sheet_name="Exchange Rates", index=False, header=True)
        # writer.save()

        cash_balance_projection_data = [
            {
                'Currency': input_data.get('currency'), 
                'Exchange Rates': input_data.get('exchange_rates'),
                'Cash - Current & PreBorrowing': input_data.get('Cash - Current & PreBorrowing'),
                'Borrowing': input_data.get('borrowing'),
                'Additional Expences 1': input_data.get('additional_expenses_1'),
                'Additional Expences 2': input_data.get('additional_expenses_2'),
                'Additional Expences 3': input_data.get('additional_expenses_3')
            } for input_data in base_data_other_info.other_info_list.get('other_sheet')
            if input_data
        ]
        cash_balance_projection_df = pd.DataFrame(cash_balance_projection_data)
        cash_balance_projection_df.to_excel(writer, sheet_name="Cash Balance Projections", index=False, header=True)
        writer.save()
        # data = {'Currency': ['USD', 'CAD', 'AUD', 'EUR'], 'Exchange Rates': [1.000000, 0.695230, 0.618820, 1.035400], 'Cash - Current & PreBorrowing': [21455041.84, 216583.15, 0, 0], 'Borrowing': ['0', '', '', ''], 'Additional Expences 1': [0, 0, 0, 0], 'Additional Expences 2': [0, 0, 0, 0], 'Additional Expences 3': [0, 0, 0, 0]}
        # data = pd.DataFrame.from_dict(data)
        xl_df_map['Cash Balance Projections'] = cash_balance_projection_df

        # book = load_workbook(file_name)
        # writer = pd.ExcelWriter(file_name, engine="openpyxl")
        # writer.book = book
        # data.to_excel(writer, sheet_name="Cash Balance Projections", index=False, header=True)
        # writer.save()

        # data = {'Currency': ['USD', 'CAD', 'AUD', 'EUR'], 'Current Credit Facility Balance': [442400000, 2000000, 0, 0]}
        # data = pd.DataFrame.from_dict(data)

        credit_balance_projection_data = [
            {
                'Currency': input_data.get('currency'), 
                'Current Credit Facility Balance': input_data.get('current_credit_facility_balance')
            } for input_data in base_data_other_info.other_info_list.get('other_sheet')]
        credit_balance_projection_df = pd.DataFrame(credit_balance_projection_data)
        credit_balance_projection_df.to_excel(writer, sheet_name="Credit Balance Projection", index=False, header=True)
        writer.save()
        xl_df_map['Credit Balance Projection'] = credit_balance_projection_df

        # book = load_workbook(file_name)
        # writer = pd.ExcelWriter(file_name, engine="openpyxl")
        # writer.book = book
        # data.to_excel(writer, sheet_name="Credit Balance Projection", index=False, header=True)
        # writer.save()

        data = {'Haircut': ['', 'SD/EBITDA', 'TD/EBITDA', 'UD/EBITDA'], '20% Conc. Limit': ['Tier 1 Obligor', 5, 7, 6], 'Unnamed: 2': ['Tier 2 Obligor', 4.25, 6, 5.25], 'Unnamed: 3': ['Tier 3 Obligor', 3.75, 5, 4.5], 'Level 1 - 10% Haircut': ['Tier 1 Obligor', 5.5, 7.5, 6.5], 'Unnamed: 5': ['Tier 2 Obligor', 4.75, 6.5, 5.75], 'Unnamed: 6': ['Tier 3 Obligor', 4.25, 5.5, 5], 'Level 2 - 20% Haircut': ['Tier 1 Obligor', 5.5, 7.5, 6.5], 'Unnamed: 8': ['Tier 2 Obligor', 4.75, 6.5, 5.75], 'Unnamed: 9': ['Tier 3 Obligor', 4.25, 5.5, 5], 'Level 3 - 35% Haircut': ['Tier 1 Obligor', 5.5, 7.5, 6.5], 'Unnamed: 11': ['Tier 2 Obligor', 4.75, 6.5, 5.75], 'Unnamed: 12': ['Tier 3 Obligor', 4.25, 5.5, 5], 'Level 4 - Max Eligibility - 50% Haircut': ['Tier 1 Obligor', 5.5, 7.5, 6.5], 'Unnamed: 14': ['Tier 2 Obligor', 4.75, 6.5, 5.75], 'Unnamed: 15': ['Tier 3 Obligor', 4.25, 5.5, 5]}
        data = pd.DataFrame.from_dict(data)
        data.to_excel(writer, sheet_name="Haircut", index=False, header=True)
        writer.save()
        xl_df_map['Haircut'] = data

        # book = load_workbook(file_name)
        # writer = pd.ExcelWriter(file_name, engine="openpyxl")
        # writer.book = book
        # data.to_excel(writer, sheet_name="Haircut", index=False, header=True)
        # writer.save()

        # # data = {'Currency': ['USD', 'CAD', 'AUD', 'EUR'], 'Current Credit Facility Balance': ['442400000', '2000000', '0', '0']}
        # # data = pd.DataFrame.from_dict(data)
        industry_list.to_excel(writer, sheet_name="Industry", index=False, header=True)
        writer.save()
        xl_df_map['Industry'] = industry_list

        # book = load_workbook(file_name)
        # writer = pd.ExcelWriter(file_name, engine="openpyxl")
        # writer.book = book
        # industry_list.to_excel(writer, sheet_name="Industry", index=False, header=True)
        # writer.save()
        # print(industry_list)

        # print(xl_df_map)

        # df.to_excel("output.xlsx")

        xl_df_map = pd.read_excel(file_name, sheet_name=['Loan List', 'Inputs', 'Exchange Rates', 'Cash Balance Projections', 'Credit Balance Projection', 'Haircut', 'Industry'])
        pickled_xl_sheet_df_map = pickle.dumps(xl_df_map)

        included_excluded_assets_map = pfltDashboardService.pflt_included_excluded_assets(xl_df_map)

        # datetime object containing current date and time
        now = datetime.now()
        
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        base_data_file = BaseDataFile(
            user_id=1,
            closing_date=report_date,
            fund_type='PFLT',
            file_data=pickled_xl_sheet_df_map,
            file_name='Generated Data ' + dt_string,
            included_excluded_assets_map=json.dumps(included_excluded_assets_map),
            extracted_base_data_info_id= bdi_id
        )
        print('Generated Data ' + dt_string)
        db.session.add(base_data_file)
        db.session.commit()

        base_data_file = commonServices.get_base_data_file(
            base_data_file_id=base_data_file.id
        )
        # if base_data_file.fund_type == "PCOF":
            # return pcofDashboardService.calculate_bb(
            #     base_data_file, selected_assets, user_id
            # )
        # else:
        selected_assets = included_excluded_assets_map['included_assets']
        wb2.close()
        writer.close()
        del writer
        del wb2
        os.remove(file_name)
        bb_response = pfltDashboardService.calculate_bb(base_data_file, selected_assets, 1)
        bb_response["base_data_file_id"] = base_data_file.id
        return ServiceResponse.success(message="Successfully processed calculation.", data=bb_response)

    except Exception as e:
        print(e)
        return ServiceResponse.error(message="Something went wrong while triggering calculation")

def get_archived_file_list():

    source_files = db.session.query(SourceFiles).filter(SourceFiles.is_archived == True).order_by(SourceFiles.uploaded_at.desc()).all()
    
    list_table = {
        "columns": [{
            "key": "fund", 
            "label": "Fund",
        }, {
            "key": "file_name", 
            "label": "File Name",
        }, {
            "key": "report_date", 
            "label": "Report Date",
        }, {
            "key": "uploaded_at", 
            "label": "Uploaded at",
        }, {
            "key": "uploaded_by", 
            "label": "Uploaded by",
        }], 
        "data": []
    }

    for source_file in source_files:
        list_table["data"].append({
            "file_id": source_file.id,
            "file_name": source_file.file_name + source_file.extension, 
            "uploaded_at": source_file.uploaded_at.strftime("%Y-%m-%d"),
            "report_date": source_file.report_date.strftime("%Y-%m-%d"),
            "fund": source_file.fund_types,
            "source_file_type": source_file.file_type,
        })
        print(source_file.fund_types)
    
    return ServiceResponse.success(data=list_table)

def update_archive(list_of_ids, to_archive):
    try:
        source_file_list = []
        for file_id in list_of_ids:
            source_file = SourceFiles.query.filter_by(id=file_id).first()

            if source_file:
                source_file.is_archived = to_archive

            source_file_list.append(source_file)

        db.session.add_all(source_file_list)
        db.session.commit()
              
        return ServiceResponse.success(message = f"Files {'added to archive' if to_archive else 'removed from archive'} successfully")

    except Exception as e:
        Log.func_error(e=e)
        return ServiceResponse.error(message="Could not update the files.", status_code = 500)

def add_base_data_other_info(
        extraction_info_id,
        determination_date, 
        fund_type, 
        other_data
    ):
    try:
        existing_record = BaseDataOtherInfo.query.filter_by(extraction_info_id=extraction_info_id).first()

        if existing_record:
            existing_record.determination_date = determination_date
            existing_record.fund_type = fund_type
            existing_record.other_info_list = other_data
            existing_record.company_id = 1
        else:
            extraction_info = ExtractedBaseDataInfo.query.filter_by(id=extraction_info_id).first()
            company_id = extraction_info.company_id
            base_data_other_info =  BaseDataOtherInfo(
                extraction_info_id = extraction_info_id,
                determination_date = determination_date,
                fund_type = fund_type,
                other_info_list = other_data,
                company_id = company_id
            )
            db.session.add(base_data_other_info)
        db.session.commit()

        return ServiceResponse.success(message="Data added sucessfully")
        
    except Exception as e:
        Log.func_error(e)
        return ServiceResponse.error(message="Failed to add")

def get_base_data_other_info(extraction_info_id, fund_type):
    try:
        other_info = db.session.query(
            BaseDataOtherInfo.id,
            BaseDataOtherInfo.extraction_info_id,
            BaseDataOtherInfo.determination_date,
            BaseDataOtherInfo.other_info_list
        ).filter(BaseDataOtherInfo.extraction_info_id == extraction_info_id).first()
        res = {}
        if other_info:
            res = {
                "id": other_info.id,
                "extraction_info_id": other_info.extraction_info_id,
                "determination_date": other_info.determination_date,
                "other_data": other_info.other_info_list
            }

        return ServiceResponse.success(data = res)
    except Exception as e:
        Log.func_error(e)
        return ServiceResponse.error()
    
def update_source_file_info(source_file, isValidated=False, isExtracted=False, validation_info=[]):
    try:
        # for source_file in source_file_list:
        source_file_detail = source_file.get("source_file")
        source_file_detail.is_validated = isValidated
        source_file_detail.is_extracted = isExtracted
        source_file_detail.validation_info = validation_info
        source_file_detail.extraction_status = "Completed" if isExtracted else "Failed"
        db.session.add(source_file_detail)
        db.session.commit()
    except Exception as e:
        Log.func_error(e)
        print(f"error on line {e.__traceback__.tb_lineno} inside {__file__}")
        
def extract_validate_store_update(source_file):
    try:
        from app import app
        with app.app_context():
            # for source_file in source_files_list:
            src_file = source_file["source_file"]
            print(src_file.id)
            
            file_name = src_file.file_name
            extension = src_file.extension
            file_full_name = file_name + extension

            if extension.lower() == '.csv' and "market" in file_name.lower():
                file_bytes = source_file["file"]
                file_stream = BytesIO(file_bytes)
                
                df = pd.read_csv(file_stream)
                
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Market and Book Value')
                excel_buffer.seek(0)
                
                source_file["file"] = excel_buffer.getvalue()
                src_file.extension = '.xlsx'

            extraction_response = extract_source_file(source_file)
            
            # Restore original extension if it was CSV
            if extension.lower() == '.csv':
                src_file.extension = extension

            file_type = src_file.file_type
            engine = db.get_engine()
        
            with engine.connect() as connection:
                sheets = connection.execute(text(f"""
                    select smm."name" from sheet_metadata_master smm 
                    join file_metadata_master fmm on smm.file_id = fmm.id 
                    where fmm."type" = '{file_type}'
                """)).fetchall()

            validation_status = None


            mismatched_data = []
            for sheet_tuple in sheets:
                sheet = sheet_tuple[0]
                if sheet in extraction_response.get("data"):
                    extracted_df = extraction_response.get("data").get(sheet)
                    validation_res = validate_uploaded_file(extracted_df, sheet_name=sheet, mismatched_data=mismatched_data)
    
                    validation_status = validation_res.get('success')
                    if validation_status == False:
                        break
                    validated_df = validation_res.get('data')
                    extraction_response.get("data")[sheet] = validated_df
                else:
                    mismatched_data.append({'sheet_name': sheet, 'is_sheet_available': False})
                        
            is_Extracted = False
            is_validated = (not bool(len(mismatched_data))) and validation_status 
            # if is_validated:
            extracted_data = extraction_response.get("data")
            store_response = store_sheet_data(data_dict=extracted_data)
            is_Extracted = store_response.get("success")

            update_source_file_info(source_file, isValidated= is_validated, isExtracted=is_Extracted, validation_info=mismatched_data)
        print(f"{file_full_name} extracted, validated, stored")
        return ServiceResponse.success()

    except Exception as e:
        # Log.func_error(e)
        print(str(e)[:200])
        return ServiceResponse.error()

def get_unmapped_cash_sec():
    engine = db.get_engine()
    with engine.connect() as connection:
        unmapped_securities = pd.DataFrame(connection.execute(text('''select 
	        distinct "Security/Facility Name" as cashfile_securities, 
	        "Security ID" as "security_id", 
	        "Issuer/Borrower Name" as "issuer_borrower_name",
	        sum("P. Lot Current Par Amount (Deal Currency)"::float) as "par_amout_deal"
        from sf_sheet_us_bank_holdings pubh 
        left join pflt_security_mapping psm on psm.cashfile_security_name = pubh."Security/Facility Name" 
        where psm.id is null group by "Security/Facility Name", "Security ID", "Issuer/Borrower Name"''')).fetchall())

    columns = [
        {"key": "security_id", "label": "Security ID", 'isEditable': False, 'isRequired': True},
        {"key": "cashfile_securities", "label": "Security/Facility Name", 'isEditable': False, 'isRequired': True},
        {"key": "issuer_borrower_name", "label": "Issuer/Borrower Name", 'isEditable': False, 'isRequired': True},
        {"key": "par_amout_deal", "label": "P. Lot Current Par Amount (Deal Currency)", 'isEditable': False, 'isRequired': True},
        # {"key": "pik_loan", "label": "PIK Loan", 'isEditable': False, 'isRequired': True}
    ]

    unmapped_securities = unmapped_securities.fillna("")
    unmapped_securities_dict = unmapped_securities.to_dict(orient='records')

    unmapped_securities_list = {
        "columns": columns,
        "data": unmapped_securities_dict
    }
    return ServiceResponse.success(data=unmapped_securities_list, message="Unmapped Securities")

def get_unmapped_pflt_sec(cash_file_security):
    try:
        cash_file_security_initial = cash_file_security.split(" ")[0]
        engine = db.get_engine()
        with engine.connect() as connection:
            matched_pflt_sec_mapping_df = pd.DataFrame(connection.execute(text(f'''
                select 
                    id, 
                    soi_name, 
                    master_comp_security_name, 
                    family_name, security_type, 
                    cashfile_security_name 
                from pflt_security_mapping 
                where company_id = 1 and
                pflt_security_mapping.cashfile_security_name is null and
                (soi_name ilike '{cash_file_security_initial}%' or family_name ilike '{cash_file_security_initial}%') 
                order by soi_name asc;''')).fetchall())

        columns_data = [
            {
                'key': "soi_name",
                'label': "SOI Name",
                'isEditable': False,
                'isRequired': True
            }, {
                'key': "master_comp_security_name",
                'label': "Master Comp Security Name",
                'isEditable': False
            },  {
                'key': "family_name",
                'label': "Family Name",
                'isEditable': True
            }, {
                'key': "security_type",
                'label': "Security Type",
                'isEditable': False
            # }, {
            #     'key': "cashfile_security_name",
            #     'label': "Cash File Security Name",
            #     'isEditable': True
            }
        ]
        pflt_sec_mapping_table = {
            "columns": columns_data,
            "data": []
        }
        matched_pflt_sec_mapping_df.columns = matched_pflt_sec_mapping_df.columns.str.replace(" ", "_")
        matched_pflt_sec_mapping_df = matched_pflt_sec_mapping_df.replace({np.nan: None})
        df_dict = matched_pflt_sec_mapping_df.to_dict(orient='records')
        pflt_sec_mapping_table["data"] = df_dict

        return ServiceResponse.success(data=pflt_sec_mapping_table, message="Probable securities")
    except Exception as e:
        Log.func_error(e=e)
        return ServiceResponse.error(message=f"Error occurred while retrieving unmapped cash file securities")
    
def get_cash_sec(security_type):
    engine = db.get_engine()
    with engine.connect() as connection:
        all_securities = pd.DataFrame(connection.execute(text('''select id, soi_name, master_comp_security_name, family_name, security_type, cashfile_security_name from pflt_security_mapping where company_id = 1 order by cashfile_security_name, master_comp_security_name ASC''')).fetchall())
        unmapped_securities = pd.DataFrame(connection.execute(text('''select
            distinct "Security/Facility Name" as cashfile_securities,
            "Security ID" as "security_id",
            "LoanX ID" as "loanx_id",
            "Issuer/Borrower Name" as "issuer_borrower_name",
            sum("P. Lot Current Par Amount (Deal Currency)"::float) as "par_amout_deal",
            "Facility Category Desc" as facility_category_desc
        from sf_sheet_us_bank_holdings pubh
        left join pflt_security_mapping psm on psm.cashfile_security_name = pubh."Security/Facility Name"
        where psm.id is null
        group by "Security/Facility Name", "Security ID", "LoanX ID", "Issuer/Borrower Name", "Facility Category Desc"''')).fetchall())
 
    unammped_securities_count = unmapped_securities.shape[0]
    all_securities_count = all_securities.shape[0]
 
    if security_type == "all":
        securities = all_securities
        columns = [{
                'key': "soi_name",
                'label': "SOI Name",
                'isEditable': False,
                'isRequired': True
            }, {
                'key': "master_comp_security_name",
                'label': "Security Name [Mastercomp file -> SOI Mapping]",
                'isEditable': False
            }, {
                'key': "family_name",
                'label': "Family Name",
                'isEditable': True
            }, {
                'key': "security_type",
                'label': "Security Type",
                'isEditable': False
            }, {
                'key': "cashfile_security_name",
                'label': "Security/Facility Name [Cashfile -> US Bank Holdings]",
                'isEditable': True
            }
        ]
        securities = securities.fillna("")
        securities_dict = securities.to_dict(orient='records')
    else:
        securities = unmapped_securities
        columns = [
            {"key": "security_id", "label": "Security ID", 'isEditable': False, 'isRequired': True},
            {"key": "loanx_id", "label": "LoanX ID", 'isEditable': False, 'isRequired': True},
            {"key": "cashfile_securities", "label": "Security/Facility Name [Cashfile -> US Bank Holdings]", 'isEditable': False, 'isRequired': True},
            {"key": "issuer_borrower_name", "label": "Issuer/Borrower Name", 'isEditable': False, 'isRequired': True},
            {"key": "par_amout_deal", "label": "P. Lot Current Par Amount (Deal Currency)", 'isEditable': False, 'isRequired': True},
            {"key": "facility_category_desc", "label": "Facility Category Desc", 'isEditable': False, 'isRequired': True}
        ]
        securities = securities.fillna("")
        securities_dict = securities.to_dict(orient='records')
        for s in securities_dict:
            if isinstance(s['par_amout_deal'], (int, float, complex)):
                s['par_amout_deal'] = numerize.numerize(float(s['par_amout_deal']), 2)
 
    unmapped_securities_list = {
        "columns": columns,
        "data": securities_dict,
        "all_securities_count": all_securities_count,
        "unmapped_securities_count": unammped_securities_count
    }
    return ServiceResponse.success(data=unmapped_securities_list, message="All Securities")

def validate_add_securities(sheet_data, fund_type):
    try:
        engine = db.get_engine()

        with engine.connect() as connection:
                sheets_info_list = connection.execute(text(f"""
                    select smm."lookup", smm.data_format from file_metadata_master fmm join sheet_metadata_master smm on fmm.id = smm.file_id 	
                    where fmm."type" = 'addbasedata' and smm.fund_id = (select id from fund where fund_name = '{fund_type}')
                """)).fetchall()

        mismatched_data = []

        for sheet_info in sheets_info_list:
            sheet_name = sheet_info[0]

            sheet_df = pd.DataFrame(sheet_data)
            sheet_df = sheet_df.drop(columns=[col for col in ['id', 'action'] if col in sheet_df.columns])

            validation_response = validate_tabular_sheet(sheet_df, sheet_name, mismatched_data, base_data_add=True)

            validation_status = validation_response.get('success')
            if validation_status == False:
                return ServiceResponse.info(success=False, data=mismatched_data)
        
        return ServiceResponse.info(data=mismatched_data)       

    except Exception as e:
        print(e)
        return ServiceResponse.error()


def add_to_base_data_table(records, fund_type, base_data_info_id, company_id, report_date):
    try:
        if fund_type == 'PCOF':
            sheet_name = "PL BB Build"
            bd_table = PcofBaseData
        elif fund_type == 'PFLT':
            sheet_name = "Loan List"
            bd_table = PfltBaseData
        elif fund_type == 'PSSL':
            sheet_name = "Portfolio"
            bd_table = PsslBaseData
        else:
            return ServiceResponse.error(message="Invalid fund type.")

        # Fetch column mapping
        base_data_mapping = {
            mapping.bd_column_name: {
                'bd_column_lookup': mapping.bd_column_lookup,
                'bd_column_datatype': mapping.bd_column_datatype
            }
            for mapping in BaseDataMapping.query.filter(
                BaseDataMapping.fund_type == fund_type,
                BaseDataMapping.bd_sheet_name == sheet_name,
            ).all()
        }

        for record in records:
            if record.get('action') == 'add':
                bd_table_obj = bd_table()
            else:
                id = record.get('id')
                bd_table_obj = bd_table.query.filter_by(id=id).first()

                if not bd_table_obj:
                    print(f"No record found with ID {id}")
                    continue

            for bd_column_name, value in record.items():
                if bd_column_name not in ['id', 'action']:
                    bd_column_lookup = base_data_mapping.get(bd_column_name).get('bd_column_lookup')
                    bd_column_datatype = base_data_mapping.get(bd_column_name).get('bd_column_datatype')
                    if bd_column_datatype == "datetime":
                        if value == '' or value is None:
                            value = None
                        else:
                            value = datetime.strptime(value, "%m-%d-%Y")
                    if bd_column_lookup:
                        setattr(bd_table_obj, bd_column_lookup, value)
                    else:
                        print(f"Warning: No mapping found for {bd_column_name}")

            
            setattr(bd_table_obj, 'base_data_info_id', base_data_info_id)
            setattr(bd_table_obj, 'company_id', company_id)
            setattr(bd_table_obj, 'report_date', report_date)

            if record.get('action') == 'add':
                setattr(bd_table_obj, 'is_manually_added', True)
                db.session.add(bd_table_obj)

        db.session.commit()

        return ServiceResponse.success(message="Data added successfully.")

    except Exception as e:
        import traceback
        print(f"Error: {e}")
        traceback.print_exc()
        return ServiceResponse.error(message="Something went wrong.")
    
def validate_uploaded_file(sheet_df, sheet_name, mismatched_data):
    try:
        engine = db.get_engine()
        
        with engine.connect() as connection:
            columns_tuple = connection.execute(text(f"""
                select cmm.column_name, cmm.is_required, cmm.data_type, cmm.column_categories, cmm.exceptions
                from column_metadata_master cmm 
                join sheet_metadata_master smm on cmm.sheet_id = smm.smm_id 
                where smm."name" = '{sheet_name}'
            """)).fetchall()
        
        # columns = [column[0] for column in columns_tuple]
        column_names = ['source_file_id']

        for column_tuple in columns_tuple:
            column = column_tuple[0]
            expected_type = column_tuple[2]
            exceptions = column_tuple[4]
            if exceptions is None:
                exceptions = []

            column_categories = column_tuple[3]
            full_column_name = column_categories + ' ' + column if column_categories is not None else column
            # print(sheet_name+ ': ' + full_column_name)
            if full_column_name not in sheet_df.columns:
                mismatched_data.append({
                    'sheet_name': sheet_name,
                    'column_name': column,
                    'is_column_available': False
                })
            else:
                column_list = sheet_df[full_column_name].tolist()
                for index in range(len(column_list)):
                    value = column_list[index]
                    if not check_data_type(value, expected_type, exceptions):
                        mismatched_data.append({
                            'sheet_name': sheet_name,
                            'column_name': column,
                            'value': value,
                            'expected_type': expected_type,
                            'actual_type': type(value).__name__,
                            'is_sheet_available': True,
                            'is_column_available': True,
                            'index': index,
                            'column_categories': column_categories
                        })

                column_names.append(full_column_name)
        
        # store only those columns that are in column metadata table
        # sheet_df = sheet_df[column_names]
        
        return ServiceResponse.success(data=sheet_df)       

    except Exception as e:
        print(str(e)[:200])
        return ServiceResponse.error()
    

def validate_key_value_sheet(sheet_data, sheet_name, mismatched_data):
    try:
        engine = db.get_engine()
        with engine.connect() as connection:
            key_info_list = connection.execute(text(f"""
                select cmm.column_lookup, cmm.is_required, cmm.data_type
                from column_metadata_master cmm 
                join sheet_metadata_master smm on cmm.sheet_id = smm.smm_id 
                where smm."lookup" = '{sheet_name}'
            """)).fetchall()

            for key in key_info_list:
                key_name = key[0]
                key_isRequired = key[1]
                value_data_type = key[2]

                if key_name not in sheet_data and key_isRequired:
                    mismatched_data.append({
                        'sheet_name': sheet_name,
                        'column_name': key_name,
                        'is_column_available': False
                    })
                else:
                    value = sheet_data[key_name]
                    if not check_value_data_type(value, value_data_type):
                        mismatched_data.append({
                            'sheet_name': sheet_name,
                            'column_name': key_name,
                            'value': value,
                            'expected_type': value_data_type,
                            'actual_type': infer_data_type(value),
                            'is_sheet_available': True,
                            'is_column_available': True,
                            'index': None,
                            'column_categories': None
                        })

        return ServiceResponse.success()
    
    except Exception as e:
        print(e)
        return ServiceResponse.error()
    
def validate_tabular_sheet(sheet_df, sheet_name, mismatched_data, base_data_add = False):
    try:
        engine = db.get_engine()
        
        with engine.connect() as connection:
            columns_tuple = connection.execute(text(f"""
                select cmm.column_lookup, cmm.column_name, cmm.is_required, cmm.data_type
                from column_metadata_master cmm 
                join sheet_metadata_master smm on cmm.sheet_id = smm.smm_id 
                where smm."lookup" = '{sheet_name}'
            """)).fetchall()
        
        for column_tuple in columns_tuple:
            column = column_tuple[1] if base_data_add else column_tuple[0]
            column_isRequired = column_tuple[2]
            expected_type = column_tuple[3]
            if column not in sheet_df.columns and column_isRequired:
                mismatched_data.append({
                    'sheet_name': sheet_name,
                    'column_name': column,
                    'is_column_available': False
                })
            else:
                column_list = sheet_df[column].tolist() if column in sheet_df.columns else []
                for index in range(len(column_list)):
                    value = column_list[index]
                    if not check_value_data_type(value, expected_type):
                        mismatched_data.append({
                            'sheet_name': sheet_name,
                            'column_name': column,
                            'value': value,
                            'expected_type': expected_type,
                            'actual_type': infer_data_type(value),
                            'is_sheet_available': True,
                            'is_column_available': True,
                            'index': index,
                            'column_categories': None
                        })
        return ServiceResponse.success(data=sheet_df)       

    except Exception as e:
        print(e)
        return ServiceResponse.error()
    

def validate_other_info_sheet(fund_type, other_data):
    try:
        engine = db.get_engine()

        with engine.connect() as connection:
                sheets_info_list = connection.execute(text(f"""
                    select smm."lookup", smm.data_format from file_metadata_master fmm join sheet_metadata_master smm on fmm.id = smm.file_id 	
                    where fmm."type" = 'otherinfo' and smm.fund_id = (select id from fund where fund_name = '{fund_type}')
                """)).fetchall()

        mismatched_data = []

        for sheet_info in sheets_info_list:
            sheet_name = sheet_info[0]
            sheet_data_format = sheet_info[1]
            sheet_data = other_data[sheet_name]
            if sheet_data_format == 'tabular':
                sheet_df = pd.DataFrame(sheet_data)
                validation_response = validate_tabular_sheet(sheet_df, sheet_name, mismatched_data)

            else:
                validation_response = validate_key_value_sheet(sheet_data, sheet_name, mismatched_data)

            validation_status = validation_response.get('success')
            if validation_status == False:
                return ServiceResponse.info(success=False, data=mismatched_data)
        
        return ServiceResponse.info(data=mismatched_data)       

    except Exception as e:
        print(e)
        return ServiceResponse.error()
    
def compare_columns(file, fund_type, company_id, file_type):
    try:
        
        match fund_type:
            case "PCOF":
                sheet_name = "PL BB Build"
                fund_id = 1
            case "PFLT":
                sheet_name = "Loan List"
                fund_id = 2
            case "PSSL":
                sheet_name = "Portfolio"
                fund_id = 3

        engine = db.get_engine()

        with engine.connect() as connection:
            # action -> addsecurit. (type of FMM), fund, comapnyId, column_id need to add
            columns_tuple = connection.execute(text(f"""
                select cmm.column_name, cmm.column_aliases, cmm.cmm_id
                from column_metadata_master cmm 
                join sheet_metadata_master smm on cmm.sheet_id = smm.smm_id 
                join file_metadata_master fmm on fmm.id = smm.file_id 
                where smm."name" = '{sheet_name}' and fmm.company_id = {company_id} and fmm.type = '{file_type}' and cmm.fund_id = {fund_id}
            """)).fetchall()

        df = pd.read_excel(file)
        file_columns = set(col.strip() for col in df.columns)

        db_column_names = set()
        db_aliases = set()
        colname_to_cmmid = dict()
        colname_to_aliases = dict()

        for col_name, aliases, cmm_id in columns_tuple:
            col_name = col_name.strip()
            db_column_names.add(col_name)
            colname_to_cmmid[col_name] = cmm_id

            alias_set = set()
            if aliases:
                for alias in aliases:
                    alias = alias.strip()
                    if alias:
                        db_aliases.add(alias)
                        alias_set.add(alias)
                        colname_to_cmmid[alias] = cmm_id

            colname_to_aliases[col_name] = alias_set

        all_known_columns = db_column_names.union(db_aliases)
        extra_in_file = [col for col in file_columns if col not in all_known_columns]
        
        file_columns_set = set(file_columns)
        missing_columns = []

        for col_name in db_column_names:
            aliases = colname_to_aliases.get(col_name, set())
            if col_name not in file_columns_set and not aliases.intersection(file_columns_set):
                missing_columns.append(col_name)

        missing_columns = sorted(missing_columns)
        extra_in_file = sorted(extra_in_file)
        
        missing_in_file = [
            {"column_name": col, "id": colname_to_cmmid.get(col)}
            for col in missing_columns
        ]

        data = {
            "missing_columns_in_file": missing_in_file,
            "extra_columns_in_file": extra_in_file
        }

        return ServiceResponse.success(data=data)
    
    except Exception as e:
        print(e)
        return ServiceResponse.error()
    

def save_columns(ids_list, mapped_columns):
    try: 
        engine = db.get_engine()
        with engine.connect() as connection:
            result = connection.execute(
                text("""
                    SELECT cmm_id, column_aliases 
                    FROM column_metadata_master 
                    WHERE cmm_id IN :ids_list
                """).bindparams(bindparam("ids_list", expanding=True)),
                {"ids_list": ids_list}
            ).fetchall()

            alias_dict = {row[0]: (None if row[1] == ['NULL'] else row[1]) for row in result}

            updates_column_aliases = [] 
            for updated_column in mapped_columns:
                cmm_id = updated_column.get('id')
                column_name = updated_column.get('column_name')
                existing_aliases = alias_dict.get(cmm_id)
                
                if not existing_aliases:
                    alias_set = set()
                else:
                    alias_set = set(existing_aliases)
                
                alias_set.add(column_name)
                new_aliases = list(alias_set)

                updates_column_aliases.append({
                    "cmm_id": cmm_id,
                    "new_aliases": new_aliases
                })

            for column_aliases in updates_column_aliases:
                connection.execute(
                    text("""
                        UPDATE column_metadata_master
                        SET column_aliases = :new_aliases
                        WHERE cmm_id = :cmm_id
                    """),
                    {"new_aliases": column_aliases["new_aliases"], "cmm_id": column_aliases["cmm_id"]}
                )
            
            connection.commit()

        return ServiceResponse.success()
    
    except Exception as e:
        print(e)
        return ServiceResponse.error()

def add_vae_data(vae_data):
    try:
        vae_data_obj = VaeData(obligor=vae_data['obligor'],
            event_type=vae_data['eventType'],
            material_modification=vae_data['materialModification'],
            vae_decision_date=vae_data['vaeDecisionDate'],
            financials_date=vae_data['financialsDate'],
            ttm_ebitda=vae_data['ttmEbitda'],
            senior_debt=vae_data['seniorDebt'],
            total_debt=vae_data['totalDebt'],
            unrestricted_cash=vae_data['unrestrictedCash'],
            net_senior_leverage=vae_data['netSeniorLeverage'],
            net_total_leverage=vae_data['netTotalLeverage'],
            interest_coverage=vae_data['interestCoverage'],
            recurring_revenue=vae_data['recurringRevenue'],
            debt_to_recurring_revenue_ratio=vae_data['debtToRecurringRevenueRatio'],
            liquidity=vae_data['liquidity'],
            assigned_value=vae_data['assignedValue'],
            created_by=1)

        db.session.add(vae_data_obj)
        db.session.commit()
        
        return ServiceResponse.success(message="VAE data added successfully")
    except Exception as e:
        raise Exception(e)

def get_vae_data():
    try:
        vae_data = VaeData.query.order_by(VaeData.vd_id.desc()).all()
        result = []
        for data in vae_data:
            temp = {
                'obligor': data.obligor,
                'eventType': data.event_type,
                'materialModification': data.material_modification,
                'vaeDecisionDate': data.vae_decision_date,
                'financialsDate': data.financials_date,
                'ttmEbitda': data.ttm_ebitda,
                'seniorDebt': data.senior_debt,
                'totalDebt': data.total_debt,
                'unrestrictedCash': data.unrestricted_cash,
                'netSeniorLeverage': data.net_senior_leverage,
                'netTotalLeverage': data.net_total_leverage,
                'interestCoverage': data.interest_coverage,
                'recurringRevenue': data.recurring_revenue,
                'debtToRecurringRevenueRatio': data.debt_to_recurring_revenue_ratio,
                'liquidity': data.liquidity,
                'assignedValue': str(data.assigned_value * 100)+'%' if data.assigned_value is not None else None
            }
            result.append(temp)
        return ServiceResponse.success(message="Success", data=result)
    except Exception as e:
        raise Exception(e)
    
def get_unmapped_securities(file_ids, fund_type):
    try:
        engine = db.get_engine()
        query = ''
        if fund_type == 'PCOF':
            query = """select distinct ssm."Issuer_Name" as asset, ssm."Asset_Name" as type from source_files sf left join sf_sheet_marketbook_1 ssm on ssm.source_file_id = sf.id left join pflt_security_mapping psm on TRIM(psm.marketvalue_issuer) = TRIM(ssm."Issuer_Name") and TRIM(psm.marketvalue_asset) = TRIM(ssm."Asset_Name") where sf.id = ANY(:file_ids) and psm.id is null and file_type = 'market_book_file'"""
        else:
            query = """SELECT DISTINCT ssubh."Security/Facility Name" as asset
                FROM source_files sf
                LEFT JOIN sf_sheet_us_bank_holdings ssubh ON ssubh.source_file_id = sf.id
                LEFT JOIN pflt_security_mapping psm 
                    ON psm.cashfile_security_name = ssubh."Security/Facility Name"
                WHERE sf.id = ANY(:file_ids)
                AND psm.id IS NULL
                AND file_type = 'cashfile'"""
        with engine.connect() as connection:
            rows = connection.execute(text(query), {"file_ids": file_ids}).fetchall()
            result = [dict(row._mapping) for row in rows]
        response = {
            'data': result,
            'count': len(result)
        }
        if len(result) > 0:
            message = f"{len(result)} unmapped securities in the selected files. Please review them in the Security mapping screen before proceeding."
            if fund_type == 'PCOF':
                message = f"{len(result)} unmapped assets from market value file in the selected files. Please review them before proceeding."
            return ServiceResponse.error(data=response, message=message, status_code=409)
        return ServiceResponse.success()
    
    except Exception as e:
        raise Exception(e)


def get_mapping_suggestions():
    family_name_rows = PfltSecurityMapping.query.filter_by(cashfile_security_name = None).distinct().all()
    family_names = [row.family_name for row in family_name_rows]
    client = OpenAIClient()
    profile = "Berwick Industrial Park"
    prompt = (
        "Given the following company profile, select the best matching company name from the list. "
        "Provide the best match and a short reasoning.\n"
        f"Profile: {profile}\n"
        f"Company Names: {family_names}\n"
        "Respond in JSON format as {\"best_match\": <company_name>, \"reasoning\": <reason>, \"all_scores\": [<company_name>: <score>, ...]}"
    )
    response = client.chat_completion([
        {"role": "system", "content": "You are a corporate analyst. Summarize company data from JSON."},
        {"role": "user", "content": prompt}
    ])
    print(response)
    clean_str = response["content"].strip('"')
    clean_str = response["content"].strip('`')
    clean_str = re.sub(r'^json\s*', '', clean_str).strip()
    parsed_json = json.loads(clean_str)
    # Try to parse the response as JSON
    try:
        result = json.loads(response["content"] if isinstance(response, dict) else response)
    except Exception:
        result = {"raw_response": response}
    return result
    